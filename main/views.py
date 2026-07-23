from datetime import datetime, timedelta, date
import csv
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth import update_session_auth_hash, login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm, UserCreationForm, AuthenticationForm
from django.contrib.admin.views.decorators import staff_member_required
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count, Avg
from django.http import HttpResponse, HttpResponseForbidden, HttpResponseRedirect
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.db.models import Sum, Q, Count, DecimalField, F, Value
from django.db.models.functions import Coalesce
from django.views.decorators.csrf import csrf_exempt
from decimal import Decimal
from datetime import datetime, timedelta, date
from django.utils import timezone
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect, Http404
from django.conf import settings
from django.urls import reverse
from django.template.loader import render_to_string
import json

from .forms import (
    XodimForm, BonusRecordForm, JarimaRecordForm,
    ProductForm, OrderRejectForm,
)
from .models import (
    Xodim, BonusRecord, JarimaRecord,
    BonusSabab, JarimaSabab, OzgartirishTarixi,
    Category, Product, ProductOrder, PointTransaction,
    Notification, PushSubscription
)
from .services import (
    get_available_shop_ball, purchase_product,
    approve_order, reject_order,
    send_notification, send_notification_to_admins
)
from .forms import (
    XodimForm, XodimTahrirlashForm, XodimRasmForm,
    BonusRecordForm, JarimaRecordForm, UserEditForm,
    OrderRejectForm
)
from .services import (
    get_available_shop_ball, purchase_product,
    approve_order, reject_order,
    send_notification, send_notification_to_admins
)


# ============================================================
# YORDAMCHI FUNKSIYALAR
# ============================================================

def oy_oraligi(yil, oy):
    boshi = date(yil, oy, 1)
    if oy == 12:
        oxiri = date(yil + 1, 1, 1) - timedelta(days=1)
    else:
        oxiri = date(yil, oy + 1, 1) - timedelta(days=1)
    return boshi, oxiri


OYLAR = {
    1: 'Yanvar', 2: 'Fevral', 3: 'Mart', 4: 'Aprel',
    5: 'May', 6: 'Iyun', 7: 'Iyul', 8: 'Avgust',
    9: 'Sentabr', 10: 'Oktabr', 11: 'Noyabr', 12: 'Dekabr'
}

OYLAR_LIST = [{'value': k, 'nom': v} for k, v in OYLAR.items()]


def hisobot_data_yig(oy_boshi, oy_oxiri, filtrlar=None):
    if filtrlar is None:
        filtrlar = {}

    xodimlar = Xodim.objects.filter(active=True, is_archived=False).order_by('-reyting_ball')

    bonus_map = {
        row['xodim']: row
        for row in BonusRecord.objects.filter(
            sana__date__gte=oy_boshi,
            sana__date__lte=oy_oxiri
        ).values('xodim').annotate(
            jami_ball=Sum('ball_miqdori'),
            jami_pul=Sum('pul_miqdori')
        )
    }

    jarima_map = {
        row['xodim']: row
        for row in JarimaRecord.objects.filter(
            sana__date__gte=oy_boshi,
            sana__date__lte=oy_oxiri
        ).values('xodim').annotate(
            jami_ball=Sum('ball_miqdori'),
            jami_pul=Sum('pul_miqdori')
        )
    }

    hisobot_data = []
    jami_bonus_ball = 0
    jami_bonus_pul = Decimal('0')
    jami_jarima_ball = 0
    jami_jarima_pul = Decimal('0')

    for xodim in xodimlar:
        b = bonus_map.get(xodim.pk, {})
        j = jarima_map.get(xodim.pk, {})

        bonus_ball = b.get('jami_ball') or 0
        bonus_pul = b.get('jami_pul') or Decimal('0')
        jarima_ball = j.get('jami_ball') or 0
        jarima_pul = j.get('jami_pul') or Decimal('0')

        bonus_filtri = filtrlar.get('bonus', '')
        jarima_filtri = filtrlar.get('jarima', '')
        if bonus_filtri == 'olgan' and bonus_ball == 0:
            continue
        if bonus_filtri == 'olmagan' and bonus_ball > 0:
            continue
        if jarima_filtri == 'olgan' and jarima_ball == 0:
            continue
        if jarima_filtri == 'olmagan' and jarima_ball > 0:
            continue

        jami_ball = bonus_ball - jarima_ball
        jami_pul = float(bonus_pul) - float(jarima_pul)

        jami_bonus_ball += bonus_ball
        jami_bonus_pul += bonus_pul
        jami_jarima_ball += jarima_ball
        jami_jarima_pul += jarima_pul

        hisobot_data.append({
            'xodim': xodim,
            'bonus_ball': bonus_ball,
            'bonus_pul': float(bonus_pul),
            'jarima_ball': jarima_ball,
            'jarima_pul': float(jarima_pul),
            'jami_ball': jami_ball,
            'jami_pul': jami_pul,
        })

    hisobot_data.sort(key=lambda x: (-x['jami_ball'], -x['bonus_ball'], x['jarima_ball']))

    jami = {
        'bonus_ball': jami_bonus_ball,
        'bonus_pul': float(jami_bonus_pul),
        'jarima_ball': jami_jarima_ball,
        'jarima_pul': float(jami_jarima_pul),
        'ball': jami_bonus_ball - jami_jarima_ball,
        'pul': float(jami_bonus_pul) - float(jami_jarima_pul),
    }
    return hisobot_data, jami


# ============================================================
# AUTHENTICATION
# ============================================================

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f"Xush kelibsiz, {username}!")
                return redirect('dashboard')
        messages.error(request, "Login yoki parol xato!")
    else:
        form = AuthenticationForm()
    
    return render(request, 'main/login.html', {'form': form})


def logout_view(request):
    logout(request)
    messages.success(request, "Tizimdan chiqdingiz!")
    return redirect('login')


def register_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, "Ro'yxatdan o'tdingiz!")
            return redirect('dashboard')
    else:
        form = UserCreationForm()
    
    return render(request, 'main/register.html', {'form': form})


# ============================================================
# DASHBOARD
# ============================================================

@login_required
def dashboard(request):
    try:
        joriy_xodim = request.user.xodim
    except Exception:
        joriy_xodim = None

    hozir = timezone.localtime(timezone.now())
    bugun = hozir.date()

    bugun_boshi = timezone.make_aware(datetime.combine(bugun, datetime.min.time()))
    bugun_oxiri = timezone.make_aware(datetime.combine(bugun, datetime.max.time()))

    kunlik_bonuslar = BonusRecord.objects.filter(
        sana__gte=bugun_boshi, sana__lte=bugun_oxiri
    ).select_related('xodim', 'sabab').order_by('-sana')

    kunlik_jarimalar = JarimaRecord.objects.filter(
        sana__gte=bugun_boshi, sana__lte=bugun_oxiri
    ).select_related('xodim', 'sabab').order_by('-sana')

    kunlik_xaridlar = ProductOrder.objects.filter(
        created_at__gte=bugun_boshi, created_at__lte=bugun_oxiri
    ).select_related('user__xodim', 'product').order_by('-created_at')

    def harakat_dict(obj, tur):
        mahalliy = timezone.localtime(obj.sana if hasattr(obj, 'sana') else obj.created_at)
        if hasattr(obj, 'sabab') and obj.sabab:
            sabab_nomi = obj.sabab.nom
        elif hasattr(obj, 'izoh') and obj.izoh:
            sabab_nomi = obj.izoh
        else:
            sabab_nomi = tur.capitalize()
        return {
            'tur': tur,
            'xodim': obj.xodim if hasattr(obj, 'xodim') else obj.user.xodim,
            'ball': obj.ball_miqdori if hasattr(obj, 'ball_miqdori') else obj.points_spent,
            'pul': obj.pul_miqdori if hasattr(obj, 'pul_miqdori') else 0,
            'sabab': sabab_nomi if tur != 'xarid' else obj.product.name,
            'sabab_izoh': obj.izoh if hasattr(obj, 'izoh') and not getattr(obj, 'sabab', None) else None,
            'sana': mahalliy,
            'vaqt': mahalliy.strftime('%H:%M'),
        }

    kunlik_harakatlar = (
        [harakat_dict(b, 'bonus') for b in kunlik_bonuslar] +
        [harakat_dict(j, 'jarima') for j in kunlik_jarimalar] +
        [harakat_dict(x, 'xarid') for x in kunlik_xaridlar]
    )
    kunlik_harakatlar.sort(key=lambda x: x['sana'], reverse=True)

    xodimlar = Xodim.objects.filter(active=True, is_archived=False).order_by('-reyting_ball')

    return render(request, 'main/dashboard.html', {
        'joriy_xodim': joriy_xodim,
        'bugun': bugun,
        'kunlik_harakatlar': kunlik_harakatlar,
        'kunlik_harakatlar_soni': len(kunlik_harakatlar),
        'xodimlar': xodimlar,
        'xodimlar_soni': xodimlar.count(),
    })


# ============================================================
# XODIMLAR
# ============================================================

@staff_member_required
def xodim_qoshish(request):
    if request.method == 'POST':
        form = XodimForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Xodim qo'shildi!")
            return redirect('xodimlar')
    else:
        form = XodimForm()
    return render(request, 'main/xodim_form.html', {
        'form': form, 'title': "Xodim Qo'shish"
    })


@login_required
def xodimlar(request):
    qs = Xodim.objects.filter(active=True, is_archived=False).order_by('-reyting_ball')
    qidiruv = request.GET.get('qidiruv', '')
    if qidiruv:
        qs = qs.filter(
            Q(ism__icontains=qidiruv) |
            Q(familya__icontains=qidiruv) |
            Q(telefon__icontains=qidiruv)
        )

    paginator = Paginator(qs, 10)
    xodimlar_page = paginator.get_page(request.GET.get('page', 1))

    all_qs = paginator.object_list
    return render(request, 'main/xodimlar.html', {
        'xodimlar': xodimlar_page,
        'umumiy_ball': sum(x.reyting_ball for x in all_qs),
        'umumiy_bonus': sum(x.bonus_ball for x in all_qs),
        'umumiy_jarima': sum(x.jarima_ball for x in all_qs),
    })


@login_required
def xodim_detail(request, pk):
    xodim = get_object_or_404(Xodim, pk=pk)
    bonuslar = BonusRecord.objects.filter(xodim=xodim).order_by('-sana')[:30]
    jarimalar = JarimaRecord.objects.filter(xodim=xodim).order_by('-sana')[:30]
    tarixlar = OzgartirishTarixi.objects.filter(xodim=xodim).order_by('-sana')[:20]
    joylashuv = Xodim.objects.filter(reyting_ball__gt=xodim.reyting_ball, is_archived=False).count() + 1
    jami_xodimlar = Xodim.objects.filter(active=True, is_archived=False).count()
    return render(request, 'main/xodim_detail.html', {
        'xodim': xodim,
        'bonuslar': bonuslar,
        'jarimalar': jarimalar,
        'tarixlar': tarixlar,
        'joylashuv': joylashuv,          # QO'SHILDI
        'jami_xodimlar': jami_xodimlar,  # QO'SHILDI
    })


@staff_member_required
def xodim_tahrirlash(request, pk):
    xodim = get_object_or_404(Xodim, pk=pk)
    
    if request.method == 'POST':
        form = XodimTahrirlashForm(request.POST, request.FILES, instance=xodim)
        if form.is_valid():
            # Eski qiymatlarni saqlash
            eski_bonus_ball = xodim.bonus_ball
            eski_bonus_pul = xodim.bonus_pul
            eski_jarima_ball = xodim.jarima_ball
            eski_jarima_pul = xodim.jarima_pul
            eski_reyting_ball = xodim.reyting_ball
            eski_reyting_pul = xodim.reyting_pul
            eski_bonus_yechilgan = xodim.bonus_pul_yechilgan
            eski_jarima_yechilgan = xodim.jarima_pul_yechilgan
            
            yangilangan_xodim = form.save(commit=False)
            
            # Bonus/jarima maydonlarini qayta tiklash
            yangilangan_xodim.bonus_ball = eski_bonus_ball
            yangilangan_xodim.bonus_pul = eski_bonus_pul
            yangilangan_xodim.jarima_ball = eski_jarima_ball
            yangilangan_xodim.jarima_pul = eski_jarima_pul
            yangilangan_xodim.reyting_ball = eski_reyting_ball
            yangilangan_xodim.reyting_pul = eski_reyting_pul
            yangilangan_xodim.bonus_pul_yechilgan = eski_bonus_yechilgan
            yangilangan_xodim.jarima_pul_yechilgan = eski_jarima_yechilgan
            
            yangilangan_xodim.save()
            
            # User login/parolni yangilash
            if yangilangan_xodim.user:
                user = yangilangan_xodim.user
                yangi_username = form.cleaned_data.get('username', '').strip()
                yangi_password = form.cleaned_data.get('password', '').strip()
                yangi_password_confirm = form.cleaned_data.get('password_confirm', '').strip()
                
                if yangi_username and yangi_username != user.username:
                    user.username = yangi_username
                    user.save(update_fields=['username'])
                
                if yangi_password and yangi_password == yangi_password_confirm:
                    user.set_password(yangi_password)
                    user.save()
                    # Parol o'zgargandan keyin session yangilash kerak
                    from django.contrib.auth import update_session_auth_hash
                    update_session_auth_hash(request, user)
            
            messages.success(request, f"{yangilangan_xodim.ism} {yangilangan_xodim.familya} ma'lumotlari yangilandi!")
            return redirect('xodim_detail', pk=yangilangan_xodim.pk)
    else:
        form = XodimTahrirlashForm(instance=xodim)
    
    tarixlar = OzgartirishTarixi.objects.filter(xodim=xodim).order_by('-sana')[:20]
    return render(request, 'main/xodim_tahrirlash.html', {
        'form': form,
        'xodim': xodim,
        'tarixlar': tarixlar,
        'title': f"{xodim.ism} {xodim.familya} — Tahrirlash",
    })


@staff_member_required
def xodim_ochirish(request, pk):
    xodim = get_object_or_404(Xodim, pk=pk)
    
    if request.method == 'POST':
        xodim_ismi = f"{xodim.ism} {xodim.familya}"
        if xodim.user:
            xodim.user.delete()
        xodim.delete()
        messages.success(request, f"'{xodim_ismi}' o'chirildi!")
        return redirect('xodimlar')
    
    return redirect('xodim_detail', pk=pk)


# ============================================================
# ARXIVLASH
# ============================================================

@staff_member_required
def arxivlangan_xodimlar(request):
    qs = Xodim.objects.filter(is_archived=True).order_by('-reyting_ball')
    qidiruv = request.GET.get('qidiruv', '')
    if qidiruv:
        qs = qs.filter(
            Q(ism__icontains=qidiruv) |
            Q(familya__icontains=qidiruv) |
            Q(telefon__icontains=qidiruv)
        )

    paginator = Paginator(qs, 10)
    xodimlar_page = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'main/arxivlangan_xodimlar.html', {
        'xodimlar': xodimlar_page,
    })


@staff_member_required
def xodim_arxivlash(request, pk):
    xodim = get_object_or_404(Xodim, pk=pk)
    if request.method == 'POST':
        xodim.is_archived = True
        xodim.save(update_fields=['is_archived'])
        messages.success(request, f"'{xodim.ism} {xodim.familya}' arxivlandi!")
        return redirect('xodimlar')
    return redirect('xodim_detail', pk=pk)


@staff_member_required
def xodim_qayta_tiklash(request, pk):
    xodim = get_object_or_404(Xodim, pk=pk)
    if request.method == 'POST':
        xodim.is_archived = False
        xodim.save(update_fields=['is_archived'])
        messages.success(request, f"'{xodim.ism} {xodim.familya}' qayta tiklandi!")
        if 'arxiv' in request.META.get('HTTP_REFERER', ''):
            return redirect('arxivlangan_xodimlar')
        return redirect('xodim_detail', pk=pk)
    return redirect('xodim_detail', pk=pk)


# ============================================================
# BONUS VA JARIMA QO'SHISH
# ============================================================

@staff_member_required
def bonus_qoshish(request):
    if request.method == 'POST':
        if request.POST.get('sabab'):
            form = BonusRecordForm(request.POST)
            if form.is_valid():
                record = form.save(commit=False)
                record.created_by = request.user
                record.save()
                from .services import send_telegram_message
                xodim = record.xodim
                xodim.refresh_from_db()
                admin_name = request.user.get_full_name() or request.user.username
                now = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')
                send_telegram_message(
                    f"✅ <b>BONUS ✅</b>\n\n"
                    f"👤 Xodim: {xodim.ism} {xodim.familya}\n"
                    f"📋 Sabab: {record.sabab.nom if record.sabab else record.izoh}\n"
                    f"📊 Ball: +{record.ball_miqdori} ball\n"
                    f"💰 Pul: +{record.pul_miqdori:,.0f} so'm\n"
                    f"📝 Izoh: {record.izoh or 'Yo\'q'}\n\n"
                    f"📊 Yangi reyting: {xodim.reyting_ball} ball ({xodim.reyting_pul:,.0f} so'm)\n"
                    f"👨‍💼 Admin: {admin_name}\n"
                    f"⏱️ Vaqt: {now}",
                    thread_id=None
                )
                messages.success(request, "Bonus qo'shildi!")
                return redirect('dashboard')
            messages.error(request, 'Formada xatolik!')
        else:
            xodim_id = request.POST.get('xodim')
            try:
                pul = float(request.POST.get('manual_pul', 0) or 0)
                ball = int(request.POST.get('manual_ball', 0) or 0)
            except ValueError:
                pul, ball = 0.0, 0

            if xodim_id and (pul > 0 or ball > 0):
                xodim = get_object_or_404(Xodim, pk=xodim_id)
                sabab_nom = request.POST.get('manual_sabab_nom', "Qo'lda kiritilgan")
                izoh = request.POST.get('izoh', '')
                toliq_izoh = sabab_nom + (f". {izoh}" if izoh else '')

                BonusRecord.objects.create(
                    xodim=xodim, sabab=None,
                    pul_miqdori=Decimal(str(pul)), ball_miqdori=ball,
                    izoh=toliq_izoh, created_by=request.user
                )
                from .services import send_telegram_message
                xodim.refresh_from_db()
                admin_name = request.user.get_full_name() or request.user.username
                now = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')
                send_telegram_message(
                    f"✅ <b>BONUS ✅</b>\n\n"
                    f"👤 Xodim: {xodim.ism} {xodim.familya}\n"
                    f"📋 Sabab: {sabab_nom}\n"
                    f"📊 Ball: +{ball} ball\n"
                    f"💰 Pul: +{pul:,.0f} so'm\n"
                    f"📝 Izoh: {izoh or 'Yo\'q'}\n\n"
                    f"📊 Yangi reyting: {xodim.reyting_ball} ball ({xodim.reyting_pul:,.0f} so'm)\n"
                    f"👨‍💼 Admin: {admin_name}\n"
                    f"⏱️ Vaqt: {now}",
                    thread_id=None
                )
                messages.success(request, "Bonus qo'shildi!")
                return redirect('dashboard')
            messages.error(request, "Xodim va miqdorlarni to'g'ri kiriting!")

    sabablar = BonusSabab.objects.filter(active=True)
    return render(request, 'main/bonus_form.html', {
        'form': BonusRecordForm(), 'sabablar': sabablar,
        'title': "Bonus Qo'shish"
    })


@staff_member_required
def jarima_qoshish(request):
    if request.method == 'POST':
        if request.POST.get('sabab'):
            form = JarimaRecordForm(request.POST)
            if form.is_valid():
                record = form.save(commit=False)
                record.created_by = request.user
                record.save()
                from .services import send_telegram_message
                xodim = record.xodim
                xodim.refresh_from_db()
                admin_name = request.user.get_full_name() or request.user.username
                now = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')
                send_telegram_message(
                    f"🔴 <b>JARIMA ❗️</b>\n\n"
                    f"👤 Xodim: {xodim.ism} {xodim.familya}\n"
                    f"📋 Sabab: {record.sabab.nom if record.sabab else record.izoh}\n"
                    f"📊 Ball: -{record.ball_miqdori} ball\n"
                    f"💰 Pul: -{record.pul_miqdori:,.0f} so'm\n"
                    f"📝 Izoh: {record.izoh or 'Yo\'q'}\n\n"
                    f"📊 Yangi reyting: {xodim.reyting_ball} ball ({xodim.reyting_pul:,.0f} so'm)\n"
                    f"👨‍💼 Admin: {admin_name}\n"
                    f"⏱️ Vaqt: {now}",
                    thread_id=None
                )
                messages.success(request, "Jarima qo'shildi!")
                return redirect('dashboard')
            messages.error(request, 'Formada xatolik!')
        else:
            xodim_id = request.POST.get('xodim')
            try:
                pul = float(request.POST.get('manual_pul', 0) or 0)
                ball = int(request.POST.get('manual_ball', 0) or 0)
            except ValueError:
                pul, ball = 0.0, 0

            if xodim_id and (pul > 0 or ball > 0):
                xodim = get_object_or_404(Xodim, pk=xodim_id)
                sabab_nom = request.POST.get('manual_sabab_nom', "Qo'lda kiritilgan")
                izoh = request.POST.get('izoh', '')

                JarimaRecord.objects.create(
                    xodim=xodim, sabab=None,
                    pul_miqdori=Decimal(str(pul)), ball_miqdori=ball,
                    izoh=f"{sabab_nom}. {izoh}".strip(' .'),
                    created_by=request.user
                )
                from .services import send_telegram_message
                toliq_izoh = f"{sabab_nom}. {izoh}".strip(' .')
                xodim.refresh_from_db()
                admin_name = request.user.get_full_name() or request.user.username
                now = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')
                send_telegram_message(
                    f"🔴 <b>JARIMA ❗️</b>\n\n"
                    f"👤 Xodim: {xodim.ism} {xodim.familya}\n"
                    f"📋 Sabab: {sabab_nom}\n"
                    f"📊 Ball: -{ball} ball\n"
                    f"💰 Pul: -{pul:,.0f} so'm\n"
                    f"📝 Izoh: {izoh or 'Yo\'q'}\n\n"
                    f"📊 Yangi reyting: {xodim.reyting_ball} ball ({xodim.reyting_pul:,.0f} so'm)\n"
                    f"👨‍💼 Admin: {admin_name}\n"
                    f"⏱️ Vaqt: {now}",
                    thread_id=None
                )
                messages.success(request, "Jarima qo'shildi!")
                return redirect('dashboard')
            messages.error(request, "Xodim va miqdorlarni to'g'ri kiriting!")

    sabablar = JarimaSabab.objects.filter(active=True)
    return render(request, 'main/jarima_form.html', {
        'form': JarimaRecordForm(), 'sabablar': sabablar,
        "title": "Jarima Qo'shish"
    })


@staff_member_required
def qayta_yuborish(request):
    from .services import send_telegram_message

    hozir = timezone.localtime(timezone.now())
    bugun = hozir.date()
    bugun_boshi = timezone.make_aware(datetime.combine(bugun, datetime.min.time()))
    bugun_oxiri = timezone.make_aware(datetime.combine(bugun, datetime.max.time()))

    bonuslar = BonusRecord.objects.filter(
        sana__gte=bugun_boshi, sana__lte=bugun_oxiri
    ).select_related('xodim', 'sabab').order_by('-sana')

    jarimalar = JarimaRecord.objects.filter(
        sana__gte=bugun_boshi, sana__lte=bugun_oxiri
    ).select_related('xodim', 'sabab').order_by('-sana')

    if request.method == 'POST':
        admin_name = request.user.get_full_name() or request.user.username
        count = 0

        barcha = request.POST.get('barcha')
        if barcha:
            selected_bonus = bonuslar
            selected_jarima = jarimalar
        else:
            bonus_ids = request.POST.getlist('bonus_ids')
            jarima_ids = request.POST.getlist('jarima_ids')
            selected_bonus = bonuslar.filter(pk__in=bonus_ids) if bonus_ids else []
            selected_jarima = jarimalar.filter(pk__in=jarima_ids) if jarima_ids else []

        for b in selected_bonus:
            xodim = b.xodim
            xodim.refresh_from_db()
            now = timezone.localtime(b.sana).strftime('%Y-%m-%d %H:%M')
            send_telegram_message(
                f"✅ <b>BONUS ✅</b>\n\n"
                f"👤 Xodim: {xodim.ism} {xodim.familya}\n"
                f"📋 Sabab: {b.sabab.nom if b.sabab else b.izoh}\n"
                f"📊 Ball: +{b.ball_miqdori} ball\n"
                f"💰 Pul: +{b.pul_miqdori:,.0f} so'm\n"
                f"📝 Izoh: {b.izoh or 'Yo\'q'}\n\n"
                f"📊 Yangi reyting: {xodim.reyting_ball} ball ({xodim.reyting_pul:,.0f} so'm)\n"
                f"👨‍💼 Admin: {admin_name}\n"
                f"⏱️ Vaqt: {now}"
            )
            count += 1

        for j in selected_jarima:
            xodim = j.xodim
            xodim.refresh_from_db()
            now = timezone.localtime(j.sana).strftime('%Y-%m-%d %H:%M')
            send_telegram_message(
                f"🔴 <b>JARIMA ❗️</b>\n\n"
                f"👤 Xodim: {xodim.ism} {xodim.familya}\n"
                f"📋 Sabab: {j.sabab.nom if j.sabab else j.izoh}\n"
                f"📊 Ball: -{j.ball_miqdori} ball\n"
                f"💰 Pul: -{j.pul_miqdori:,.0f} so'm\n"
                f"📝 Izoh: {j.izoh or 'Yo\'q'}\n\n"
                f"📊 Yangi reyting: {xodim.reyting_ball} ball ({xodim.reyting_pul:,.0f} so'm)\n"
                f"👨‍💼 Admin: {admin_name}\n"
                f"⏱️ Vaqt: {now}"
            )
            count += 1

        if count > 0:
            messages.success(request, f"Telegram ga {count} ta xabar qayta yuborildi!")
        else:
            messages.warning(request, "Hech qanday yozuv tanlanmadi!")

        return redirect('admin_dashboard')

    return render(request, 'main/qayta_yuborish.html', {
        'bonuslar': bonuslar,
        'jarimalar': jarimalar,
        'bugun': bugun,
    })


# ============================================================
# BONUS VA JARIMA O'CHIRISH
# ============================================================

@staff_member_required
def bonus_ochirish(request, pk):
    bonus = get_object_or_404(BonusRecord, pk=pk)
    xodim = bonus.xodim

    if request.method == 'POST':
        sabab = request.POST.get('sabab', '').strip()
        if not sabab:
            messages.error(request, "O'chirish sababini yozishingiz kerak!")
            return redirect('bonus_ochirish', pk=bonus.pk)

        ball = bonus.ball_miqdori
        pul = bonus.pul_miqdori
        bonus.delete()

        OzgartirishTarixi.objects.create(
            xodim=xodim, admin=request.user,
            sabab=f"Bonus o'chirildi. Sabab: {sabab}",
        )
        messages.success(request, f"Bonus o'chirildi! ({ball} ball)")
        return redirect('xodim_detail', pk=xodim.pk)

    return render(request, 'main/ochirish_tasdiqlash.html', {
        'bonus': bonus, 'xodim': xodim, 'tur': 'bonus',
        'ball': bonus.ball_miqdori, 'pul': bonus.pul_miqdori, 'sana': bonus.sana,
    })


@staff_member_required
def jarima_ochirish(request, pk):
    jarima = get_object_or_404(JarimaRecord, pk=pk)
    xodim = jarima.xodim

    if request.method == 'POST':
        sabab = request.POST.get('sabab', '').strip()
        if not sabab:
            messages.error(request, "O'chirish sababini yozishingiz kerak!")
            return redirect('jarima_ochirish', pk=jarima.pk)

        ball = jarima.ball_miqdori
        pul = jarima.pul_miqdori
        jarima.delete()

        OzgartirishTarixi.objects.create(
            xodim=xodim, admin=request.user,
            sabab=f"Jarima o'chirildi. Sabab: {sabab}",
        )
        messages.success(request, f"Jarima o'chirildi! ({ball} ball)")
        return redirect('xodim_detail', pk=xodim.pk)

    return render(request, 'main/ochirish_tasdiqlash.html', {
        'jarima': jarima, 'xodim': xodim, 'tur': 'jarima',
        'ball': jarima.ball_miqdori, 'pul': jarima.pul_miqdori, 'sana': jarima.sana,
    })


# ============================================================
# PUL YECHISH (MUHIM - TUZATILGAN)
# ============================================================
# views.py - bonus_pul_yechish funksiyasini tuzatilgan versiyasi

@staff_member_required
def bonus_pul_yechish(request, pk):
    """Bonus pulini yechish"""
    xodim = get_object_or_404(Xodim, pk=pk)
    
    if request.method == 'POST':
        try:
            yechiladigan_pul = Decimal(str(float(request.POST.get('pul', 0))))
            sabab = request.POST.get('sabab', '').strip()
            
            if yechiladigan_pul <= 0:
                messages.error(request, "Pul miqdori 0 dan katta bo'lishi kerak!")
                return redirect('bonus_pul_yechish', pk=xodim.pk)
            
            if not sabab:
                messages.error(request, "Sababni yozishingiz kerak!")
                return redirect('bonus_pul_yechish', pk=xodim.pk)
            
            if yechiladigan_pul > xodim.jami_bonus_pul:
                messages.error(request, f"Yetarli bonus pul mavjud emas! Mavjud: {xodim.jami_bonus_pul:,.0f} so'm")
                return redirect('bonus_pul_yechish', pk=xodim.pk)
            
            # Eski qiymatlarni saqlash
            eski_bonus_pul = xodim.bonus_pul
            eski_bonus_yechilgan = xodim.bonus_pul_yechilgan
            
            # Yechish
            xodim.bonus_pul_yechilgan += yechiladigan_pul
            xodim.reyting_pul = xodim.jami_bonus_pul - xodim.jami_jarima_pul
            xodim.save(update_fields=['bonus_pul_yechilgan', 'reyting_pul'])
            
            # TO'G'RI - FAQAT MAVJUD MAYDONLAR
            OzgartirishTarixi.objects.create(
                xodim=xodim,
                admin=request.user,
                sabab=f"Bonus pulidan {yechiladigan_pul:,.0f} so'm yechildi. Sabab: {sabab}",
                eski_bonus_pul=eski_bonus_pul,
                yangi_bonus_pul=xodim.bonus_pul,
                # eski_bonus_yechilgan va yangi_bonus_yechilgan maydonlarini olib tashladik
            )
            
            messages.success(request, f"✅ {yechiladigan_pul:,.0f} so'm bonus pul yechildi! Qolgan: {xodim.jami_bonus_pul:,.0f} so'm")
            return redirect('xodim_detail', pk=xodim.pk)
            
        except (ValueError, TypeError) as e:
            messages.error(request, f"Xatolik: {e}")
            return redirect('bonus_pul_yechish', pk=xodim.pk)
    
    return render(request, 'main/pul_yechish.html', {
        'xodim': xodim,
        'tur': 'bonus',
        'umumiy': xodim.bonus_pul,
        'yechilgan': xodim.bonus_pul_yechilgan,
        'qoldiq': xodim.jami_bonus_pul,
    })


@staff_member_required
def jarima_pul_yechish(request, pk):
    """Jarima pulini yechish"""
    xodim = get_object_or_404(Xodim, pk=pk)
    
    if request.method == 'POST':
        try:
            yechiladigan_pul = Decimal(str(float(request.POST.get('pul', 0))))
            sabab = request.POST.get('sabab', '').strip()
            
            if yechiladigan_pul <= 0:
                messages.error(request, "Pul miqdori 0 dan katta bo'lishi kerak!")
                return redirect('jarima_pul_yechish', pk=xodim.pk)
            
            if not sabab:
                messages.error(request, "Sababni yozishingiz kerak!")
                return redirect('jarima_pul_yechish', pk=xodim.pk)
            
            if yechiladigan_pul > xodim.jami_jarima_pul:
                messages.error(request, f"Yetarli jarima pul mavjud emas! Mavjud: {xodim.jami_jarima_pul:,.0f} so'm")
                return redirect('jarima_pul_yechish', pk=xodim.pk)
            
            # Eski qiymatlarni saqlash
            eski_jarima_pul = xodim.jarima_pul
            
            # Yechish
            xodim.jarima_pul_yechilgan += yechiladigan_pul
            xodim.reyting_pul = xodim.jami_bonus_pul - xodim.jami_jarima_pul
            xodim.save(update_fields=['jarima_pul_yechilgan', 'reyting_pul'])
            
            # TO'G'RI - FAQAT MAVJUD MAYDONLAR
            OzgartirishTarixi.objects.create(
                xodim=xodim,
                admin=request.user,
                sabab=f"Jarima pulidan {yechiladigan_pul:,.0f} so'm yechildi. Sabab: {sabab}",
                eski_jarima_pul=eski_jarima_pul,
                yangi_jarima_pul=xodim.jarima_pul,
            )
            
            messages.success(request, f"✅ {yechiladigan_pul:,.0f} so'm jarima pul yechildi! Qolgan: {xodim.jami_jarima_pul:,.0f} so'm")
            return redirect('xodim_detail', pk=xodim.pk)
            
        except (ValueError, TypeError) as e:
            messages.error(request, f"Xatolik: {e}")
            return redirect('jarima_pul_yechish', pk=xodim.pk)
    
    return render(request, 'main/pul_yechish.html', {
        'xodim': xodim,
        'tur': 'jarima',
        'umumiy': xodim.jarima_pul,
        'yechilgan': xodim.jarima_pul_yechilgan,
        'qoldiq': xodim.jami_jarima_pul,
    })


@staff_member_required
def jarima_pul_yechish(request, pk):
    """Jarima pulini yechish - PUL HAQIQATAN KAMAYADI"""
    xodim = get_object_or_404(Xodim, pk=pk)
    
    if request.method == 'POST':
        try:
            yechiladigan_pul = Decimal(str(float(request.POST.get('pul', 0))))
            sabab = request.POST.get('sabab', '').strip()
            
            if yechiladigan_pul <= 0:
                messages.error(request, "Pul miqdori 0 dan katta bo'lishi kerak!")
                return redirect('jarima_pul_yechish', pk=xodim.pk)
            
            if not sabab:
                messages.error(request, "Sababni yozishingiz kerak!")
                return redirect('jarima_pul_yechish', pk=xodim.pk)
            
            jami_mavjud = xodim.jarima_pul - xodim.jarima_pul_yechilgan
            if yechiladigan_pul > jami_mavjud:
                messages.error(request, f"Yetarli jarima pul mavjud emas! Mavjud: {jami_mavjud:,.0f} so'm")
                return redirect('jarima_pul_yechish', pk=xodim.pk)
            
            # YECHISH
            xodim.jarima_pul_yechilgan += yechiladigan_pul
            xodim.reyting_pul = (xodim.bonus_pul - xodim.bonus_pul_yechilgan) - (xodim.jarima_pul - xodim.jarima_pul_yechilgan)
            xodim.save(update_fields=['jarima_pul_yechilgan', 'reyting_pul'])
            
            OzgartirishTarixi.objects.create(
                xodim=xodim,
                admin=request.user,
                sabab=f"Jarima pulidan {yechiladigan_pul:,.0f} so'm yechildi. Sabab: {sabab}",
            )
            
            qolgan_pul = xodim.jarima_pul - xodim.jarima_pul_yechilgan
            messages.success(request, f"✅ {yechiladigan_pul:,.0f} so'm jarima pul yechildi! Qolgan: {qolgan_pul:,.0f} so'm")
            return redirect('xodim_detail', pk=xodim.pk)
            
        except (ValueError, TypeError) as e:
            messages.error(request, f"Xatolik: {e}")
            return redirect('jarima_pul_yechish', pk=xodim.pk)
    
    jami_mavjud = xodim.jarima_pul - xodim.jarima_pul_yechilgan
    return render(request, 'main/pul_yechish.html', {
        'xodim': xodim,
        'tur': 'jarima',
        'umumiy': xodim.jarima_pul,
        'yechilgan': xodim.jarima_pul_yechilgan,
        'qoldiq': jami_mavjud,
    })
# main/views.py ga qo'shing

@staff_member_required
def oylik_hisobot_csv(request):
    """Oylik hisobotni CSV formatda yuklab olish"""
    oy = int(request.GET.get('oy', timezone.now().month))
    yil = int(request.GET.get('yil', timezone.now().year))
    bonus_filtri = request.GET.get('bonus_filtri', '')
    jarima_filtri = request.GET.get('jarima_filtri', '')

    oy_boshi, oy_oxiri = oy_oraligi(yil, oy)
    hisobot_data, jami = hisobot_data_yig(
        oy_boshi, oy_oxiri,
        filtrlar={'bonus': bonus_filtri, 'jarima': jarima_filtri}
    )
    oy_nomi = OYLAR.get(oy, '')

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="oylik_hisobot_{oy_nomi}_{yil}.csv"'
    response.write('\ufeff')  # BOM - Excel uchun

    writer = csv.writer(response)
    writer.writerow([f"OYLIK HISOBOT — {oy_nomi} {yil}"])
    writer.writerow([f"Hisobot davri: {oy_boshi.strftime('%d.%m.%Y')} — {oy_oxiri.strftime('%d.%m.%Y')}"])
    writer.writerow([])
    writer.writerow([
        '№', 'Xodim', 'Bonus Ball', "Bonus Pul (so'm)",
        'Jarima Ball', "Jarima Pul (so'm)",
        'Jami Ball', "Jami Pul (so'm)", 'Reyting'
    ])

    for idx, item in enumerate(hisobot_data, 1):
        writer.writerow([
            idx,
            f"{item['xodim'].ism} {item['xodim'].familya}",
            item['bonus_ball'], item['bonus_pul'],
            item['jarima_ball'], item['jarima_pul'],
            item['jami_ball'], item['jami_pul'],
            item['xodim'].reyting_ball,
        ])

    writer.writerow([])
    writer.writerow([
        'JAMI', '',
        jami['bonus_ball'], jami['bonus_pul'],
        jami['jarima_ball'], jami['jarima_pul'],
        jami['ball'], jami['pul'],
        '',
    ])
    return response


@staff_member_required
def oylik_hisobot_pdf(request):
    """Oylik hisobotni PDF formatda yuklab olish"""
    oy = int(request.GET.get('oy', timezone.now().month))
    yil = int(request.GET.get('yil', timezone.now().year))
    bonus_filtri = request.GET.get('bonus_filtri', '')
    jarima_filtri = request.GET.get('jarima_filtri', '')

    oy_boshi, oy_oxiri = oy_oraligi(yil, oy)
    hisobot_data, jami = hisobot_data_yig(
        oy_boshi, oy_oxiri,
        filtrlar={'bonus': bonus_filtri, 'jarima': jarima_filtri}
    )
    oy_nomi = OYLAR.get(oy, '')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="oylik_hisobot_{oy_nomi}_{yil}.pdf"'

    doc = SimpleDocTemplate(
        response, pagesize=landscape(A4),
        rightMargin=10, leftMargin=10, topMargin=20, bottomMargin=20
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'],
        fontSize=16, alignment=1, spaceAfter=20,
        textColor=colors.HexColor('#2001FF')
    )

    table_data = [[
        '№', 'Xodim', 'Bonus Ball', "Bonus Pul (so'm)",
        'Jarima Ball', "Jarima Pul (so'm)",
        'Jami Ball', "Jami Pul (so'm)", 'Reyting'
    ]]

    for idx, item in enumerate(hisobot_data, 1):
        table_data.append([
            str(idx),
            f"{item['xodim'].ism} {item['xodim'].familya}",
            f"+{item['bonus_ball']}",
            f"{item['bonus_pul']:,.0f}",
            f"-{item['jarima_ball']}",
            f"{item['jarima_pul']:,.0f}",
            str(item['jami_ball']),
            f"{item['jami_pul']:,.0f}",
            str(item['xodim'].reyting_ball),
        ])

    table_data.append([
        'JAMI', '',
        f"+{jami['bonus_ball']}", f"{jami['bonus_pul']:,.0f}",
        f"-{jami['jarima_ball']}", f"{jami['jarima_pul']:,.0f}",
        str(jami['ball']), f"{jami['pul']:,.0f}",
        '',
    ])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2001FF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8E8E8')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
        ('GRID', (0, -1), (-1, -1), 0.5, colors.black),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),
        ('ALIGN', (4, 1), (4, -1), 'CENTER'),
        ('ALIGN', (6, 1), (6, -1), 'CENTER'),
        ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
        ('ALIGN', (5, 1), (5, -1), 'RIGHT'),
        ('ALIGN', (7, 1), (7, -1), 'RIGHT'),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
    ]))
    table._argW = [30, 120, 50, 65, 50, 65, 50, 65, 50]

    title = Paragraph(f"<b>OYLIK HISOBOT — {oy_nomi} {yil}</b>", title_style)
    sub_title = Paragraph(
        f"<font size=9>"
        f"Hisobot davri: {oy_boshi.strftime('%d.%m.%Y')} — {oy_oxiri.strftime('%d.%m.%Y')}<br/>"
        f"Jami xodimlar: {len(hisobot_data)} ta | "
        f"Jami ball: {jami['ball']} | "
        f"Jami pul: {jami['pul']:,.0f} so'm"
        f"</font>",
        styles['Normal']
    )
    doc.build([title, Spacer(1, 10), sub_title, Spacer(1, 20), table])
    return response



@staff_member_required
def jarima_pul_yechish(request, pk):
    """Jarima pulini yechish"""
    xodim = get_object_or_404(Xodim, pk=pk)
    
    if request.method == 'POST':
        try:
            # Pul miqdorini olish
            yechiladigan_pul_str = request.POST.get('pul', '0')
            yechiladigan_pul = Decimal(str(float(yechiladigan_pul_str)))
            sabab = request.POST.get('sabab', '').strip()
            
            print(f"DEBUG: Yechiladigan pul: {yechiladigan_pul}, Sabab: {sabab}")  # Debug uchun
            
            # TEKSHIRISHLAR
            if yechiladigan_pul <= 0:
                messages.error(request, "❌ Pul miqdori 0 dan katta bo'lishi kerak!")
                return redirect('jarima_pul_yechish', pk=xodim.pk)
            
            if not sabab:
                messages.error(request, "❌ Sababni yozishingiz kerak!")
                return redirect('jarima_pul_yechish', pk=xodim.pk)
            
            # Mavjud jarima pulini hisoblash
            mavjud_jarima_pul = xodim.jarima_pul - xodim.jarima_pul_yechilgan
            print(f"DEBUG: Mavjud jarima pul: {mavjud_jarima_pul}")  # Debug uchun
            
            if yechiladigan_pul > mavjud_jarima_pul:
                messages.error(request, f"❌ Yetarli jarima pul mavjud emas! Mavjud: {mavjud_jarima_pul:,.0f} so'm")
                return redirect('jarima_pul_yechish', pk=xodim.pk)
            
            # ============================================
            # JARIMA PULINI YECHISH
            # ============================================
            
            # 1. Yechilgan pulni ko'paytirish
            xodim.jarima_pul_yechilgan += yechiladigan_pul
            
            # 2. Reyting pulni qayta hisoblash
            # Sof reyting pul = (bonus - yechilgan_bonus) - (jarima - yechilgan_jarima)
            xodim.reyting_pul = (xodim.bonus_pul - xodim.bonus_pul_yechilgan) - (xodim.jarima_pul - xodim.jarima_pul_yechilgan)
            
            # 3. Saqlash
            xodim.save(update_fields=['jarima_pul_yechilgan', 'reyting_pul'])
            
            # 4. Tarixga yozish
            OzgartirishTarixi.objects.create(
                xodim=xodim,
                admin=request.user,
                sabab=f"Jarima pulidan {yechiladigan_pul:,.0f} so'm yechildi. Sabab: {sabab}",
            )
            
            # 5. Natijani ko'rsatish
            qolgan_pul = xodim.jarima_pul - xodim.jarima_pul_yechilgan
            messages.success(
                request, 
                f"✅ {yechiladigan_pul:,.0f} so'm jarima pul yechildi!\n"
                f"💰 Umumiy jarima: {xodim.jarima_pul:,.0f} so'm\n"
                f"📤 Yechilgan: {xodim.jarima_pul_yechilgan:,.0f} so'm\n"
                f"💵 Qolgan: {qolgan_pul:,.0f} so'm"
            )
            return redirect('xodim_detail', pk=xodim.pk)
            
        except (ValueError, TypeError) as e:
            print(f"DEBUG XATO: {e}")  # Debug uchun
            messages.error(request, f"❌ Xatolik: {e}. Iltimos, to'g'ri pul miqdorini kiriting!")
            return redirect('jarima_pul_yechish', pk=xodim.pk)
        except Exception as e:
            print(f"DEBUG UMUMIY XATO: {e}")  # Debug uchun
            messages.error(request, f"❌ Kutilmagan xatolik: {e}")
            return redirect('jarima_pul_yechish', pk=xodim.pk)
    
    # GET so'rov - Formani ko'rsatish
    mavjud_pul = xodim.jarima_pul - xodim.jarima_pul_yechilgan
    context = {
        'xodim': xodim,
        'tur': 'jarima',
        'umumiy': xodim.jarima_pul,
        'yechilgan': xodim.jarima_pul_yechilgan,
        'qoldiq': mavjud_pul,
    }
    return render(request, 'main/pul_yechish.html', context)

# ============================================================
# REYTINGDAN YECHISH
# ============================================================

@staff_member_required
def reytingdan_yechish(request, pk):
    xodim = get_object_or_404(Xodim, pk=pk)

    if request.method == 'POST':
        sabab = request.POST.get('sabab', '').strip()
        if not sabab:
            messages.error(request, "Sababni yozishingiz kerak!")
            return redirect('reytingdan_yechish', pk=xodim.pk)

        try:
            ball = int(request.POST.get('ball', 0))
            pul = Decimal(str(float(request.POST.get('pul', 0))))
        except (ValueError, TypeError):
            messages.error(request, "Ball va pul to'g'ri kiriting!")
            return redirect('reytingdan_yechish', pk=xodim.pk)

        if ball == 0 and pul == 0:
            messages.error(request, "Kamida bitta qiymat kiriting!")
            return redirect('reytingdan_yechish', pk=xodim.pk)

        qayerdan = request.POST.get('qayerdan', 'reyting')

        if qayerdan == 'bonus':
            if ball > 0:
                if ball > xodim.jami_bonus_ball:
                    messages.error(request, f"Yetarli bonus ball mavjud emas! Mavjud: {xodim.jami_bonus_ball} ball")
                    return redirect('reytingdan_yechish', pk=xodim.pk)
                xodim.bonus_ball_yechilgan += ball
            elif ball < 0:
                qaytarish = abs(ball)
                if qaytarish > xodim.bonus_ball_yechilgan:
                    messages.error(request, f"Yechilgan bonus balldan ko'p qaytarib bo'lmaydi! Yechilgan: {xodim.bonus_ball_yechilgan} ball")
                    return redirect('reytingdan_yechish', pk=xodim.pk)
                xodim.bonus_ball_yechilgan -= qaytarish
            if pul > 0:
                if pul > xodim.jami_bonus_pul:
                    messages.error(request, f"Yetarli bonus pul mavjud emas! Mavjud: {xodim.jami_bonus_pul:,.0f} so'm")
                    return redirect('reytingdan_yechish', pk=xodim.pk)
                xodim.bonus_pul_yechilgan += pul
            elif pul < 0:
                qaytarish = abs(pul)
                if qaytarish > xodim.bonus_pul_yechilgan:
                    messages.error(request, f"Yechilgan bonus puldan ko'p qaytarib bo'lmaydi!")
                    return redirect('reytingdan_yechish', pk=xodim.pk)
                xodim.bonus_pul_yechilgan -= qaytarish
            izoh_tekst = "Bonusdan yechildi"
        elif qayerdan == 'jarima':
            if ball > 0:
                if ball > xodim.jami_jarima_ball:
                    messages.error(request, f"Yetarli jarima ball mavjud emas! Mavjud: {xodim.jami_jarima_ball} ball")
                    return redirect('reytingdan_yechish', pk=xodim.pk)
                xodim.jarima_ball_yechilgan += ball
            elif ball < 0:
                qaytarish = abs(ball)
                if qaytarish > xodim.jarima_ball_yechilgan:
                    messages.error(request, f"Yechilgan jarima balldan ko'p qaytarib bo'lmaydi!")
                    return redirect('reytingdan_yechish', pk=xodim.pk)
                xodim.jarima_ball_yechilgan -= qaytarish
            if pul > 0:
                if pul > xodim.jami_jarima_pul:
                    messages.error(request, f"Yetarli jarima pul mavjud emas! Mavjud: {xodim.jami_jarima_pul:,.0f} so'm")
                    return redirect('reytingdan_yechish', pk=xodim.pk)
                xodim.jarima_pul_yechilgan += pul
            elif pul < 0:
                qaytarish = abs(pul)
                if qaytarish > xodim.jarima_pul_yechilgan:
                    messages.error(request, f"Yechilgan jarima puldan ko'p qaytarib bo'lmaydi!")
                    return redirect('reytingdan_yechish', pk=xodim.pk)
                xodim.jarima_pul_yechilgan -= qaytarish
            izoh_tekst = "Jarimadan yechildi"
        else:
            xodim.reyting_ball -= ball
            xodim.reyting_pul -= pul
            izoh_tekst = "Reytingdan yechildi"

        xodim.reyting_ball = xodim.jami_bonus_ball - xodim.jami_jarima_ball
        xodim.reyting_pul = xodim.jami_bonus_pul - xodim.jami_jarima_pul
        xodim.save()

        OzgartirishTarixi.objects.create(
            xodim=xodim, admin=request.user,
            sabab=f"{izoh_tekst}: {ball} ball, {pul} so'm. Sabab: {sabab}",
        )
        messages.success(request, f"{izoh_tekst}: {ball} ball, {float(pul):,.0f} so'm!")
        return redirect('xodim_detail', pk=xodim.pk)

    return render(request, 'main/reytingdan_yechish.html', {'xodim': xodim})


# ============================================================
# REYTINGLAR
# ============================================================

@login_required
def reytinglar(request):
    davr = request.GET.get('davr', 'umumiy')
    bugun = timezone.now().date()

    hafta_boshi = bugun - timedelta(days=bugun.weekday())
    hafta_oxiri = hafta_boshi + timedelta(days=6)

    try:
        tanlangan_oy = int(request.GET.get('oy', bugun.month))
        tanlangan_yil = int(request.GET.get('yil', bugun.year))
    except (TypeError, ValueError):
        tanlangan_oy = bugun.month
        tanlangan_yil = bugun.year

    oy_nomi = OYLAR.get(tanlangan_oy, '')
    oy_boshi, oy_oxiri = oy_oraligi(tanlangan_yil, tanlangan_oy)
    yil_boshi = date(tanlangan_yil, 1, 1)
    yil_oxiri = date(tanlangan_yil, 12, 31)

    xodimlar = Xodim.objects.filter(active=True, is_archived=False)

    def ballo_map(Model, filter_kwargs):
        return {
            row['xodim']: row
            for row in Model.objects.filter(**filter_kwargs).values('xodim').annotate(
                jami_ball=Sum('ball_miqdori'),
                jami_pul=Sum('pul_miqdori')
            )
        }

    kbm = ballo_map(BonusRecord, {'sana__date': bugun})
    kjm = ballo_map(JarimaRecord, {'sana__date': bugun})
    hbm = ballo_map(BonusRecord, {'sana__date__gte': hafta_boshi, 'sana__date__lte': hafta_oxiri})
    hjm = ballo_map(JarimaRecord, {'sana__date__gte': hafta_boshi, 'sana__date__lte': hafta_oxiri})
    obm = ballo_map(BonusRecord, {'sana__date__gte': oy_boshi, 'sana__date__lte': oy_oxiri})
    ojm = ballo_map(JarimaRecord, {'sana__date__gte': oy_boshi, 'sana__date__lte': oy_oxiri})
    ybm = ballo_map(BonusRecord, {'sana__date__gte': yil_boshi, 'sana__date__lte': yil_oxiri})
    yjm = ballo_map(JarimaRecord, {'sana__date__gte': yil_boshi, 'sana__date__lte': yil_oxiri})

    def get_bb(m, pk): return m.get(pk, {}).get('jami_ball') or 0
    def get_bp(m, pk): return float(m.get(pk, {}).get('jami_pul') or 0)

    xodimlar_list = []
    for xodim in xodimlar:
        pk = xodim.pk
        d = {'xodim': xodim}

        for prefix, bm, jm in [
            ('kunlik', kbm, kjm),
            ('haftalik', hbm, hjm),
            ('oylik', obm, ojm),
            ('yillik', ybm, yjm),
        ]:
            bb = get_bb(bm, pk)
            bp = get_bp(bm, pk)
            jb = get_bb(jm, pk)
            jp = get_bp(jm, pk)
            d[f'{prefix}_bonus_ball'] = bb
            d[f'{prefix}_bonus_pul'] = bp
            d[f'{prefix}_jarima_ball'] = jb
            d[f'{prefix}_jarima_pul'] = jp
            d[f'{prefix}_reyting_ball'] = bb - jb
            d[f'{prefix}_reyting_pul'] = bp - jp

        d['umumiy_bonus_ball'] = xodim.bonus_ball
        d['umumiy_bonus_pul'] = float(xodim.bonus_pul)
        d['umumiy_jarima_ball'] = xodim.jarima_ball
        d['umumiy_jarima_pul'] = float(xodim.jarima_pul)
        d['umumiy_reyting_ball'] = xodim.reyting_ball
        d['umumiy_reyting_pul'] = float(xodim.reyting_pul)

        xodimlar_list.append(d)

    sort_key = f'{davr}_reyting_ball'
    sorted_xodimlar = sorted(
        xodimlar_list,
        key=lambda x: (
            -x.get(sort_key, 0),
            x.get(f'{davr}_jarima_ball', 0),
            -x.get(f'{davr}_bonus_ball', 0),
            x['xodim'].id
        )
    )

    jami_ball_davr = sum(x.get(sort_key, 0) for x in xodimlar_list)
    jami_pul_davr = sum(x.get(f'{davr}_reyting_pul', 0) for x in xodimlar_list)

    podium = {}
    for i, xd in enumerate(sorted_xodimlar[:3], 1):
        xodim = xd['xodim']
        podium[i] = {
            'name': f"{xodim.ism} {xodim.familya[0]}." if xodim.familya else xodim.ism,
            'full_name': f"{xodim.ism} {xodim.familya}",
            'initials': f"{xodim.ism[0]}{xodim.familya[0]}" if xodim.ism and xodim.familya else '?',
            'score': xd.get(sort_key, 0),
            'pul': xd.get(f'{davr}_reyting_pul', 0),
            'rasm': xodim.rasm.url if xodim.rasm else None,
        }

    return render(request, 'main/reytinglar.html', {
        'xodimlar': sorted_xodimlar,
        'jami_xodimlar': xodimlar.count(),
        'podium': podium,
        'davr': davr,
        'bugun': bugun,
        'hafta_boshi': hafta_boshi,
        'hafta_oxiri': hafta_oxiri,
        'oylar': OYLAR_LIST,
        'yillar': range(bugun.year - 4, bugun.year + 1),
        'tanlangan_oy': tanlangan_oy,
        'tanlangan_yil': tanlangan_yil,
        'oy_nomi': oy_nomi,
        'jami_ball_davr': jami_ball_davr,
        'jami_pul_davr': jami_pul_davr,
    })


# ============================================================
# OYLIK HISOBOT
# ============================================================

@staff_member_required
def oylik_hisobot(request):
    oy = int(request.GET.get('oy', timezone.now().month))
    yil = int(request.GET.get('yil', timezone.now().year))
    bonus_filtri = request.GET.get('bonus_filtri', '')
    jarima_filtri = request.GET.get('jarima_filtri', '')

    oy_boshi, oy_oxiri = oy_oraligi(yil, oy)
    hisobot_data, jami = hisobot_data_yig(
        oy_boshi, oy_oxiri,
        filtrlar={'bonus': bonus_filtri, 'jarima': jarima_filtri}
    )

    bugun = timezone.now().date()

    return render(request, 'main/oylik_hisobot.html', {
        'hisobot_data': hisobot_data,
        'oy_nomi': OYLAR.get(oy, ''),
        'tanlangan_oy': oy,
        'tanlangan_yil': yil,
        'oylar': OYLAR_LIST,
        'yillar': range(bugun.year - 3, bugun.year + 1),
        'oy_boshi': oy_boshi,
        'oy_oxiri': oy_oxiri,
        'jami_bonus_ball': jami['bonus_ball'],
        'jami_jarima_ball': jami['jarima_ball'],
        'jami_ball': jami['ball'],
        'jami_pul': jami['pul'],
        'xodimlar_soni': len(hisobot_data),
        'bonus_filtri': bonus_filtri,
        'jarima_filtri': jarima_filtri,
    })


@staff_member_required
def oylik_hisobot_excel(request):
    oy = int(request.GET.get('oy', timezone.now().month))
    yil = int(request.GET.get('yil', timezone.now().year))
    bonus_filtri = request.GET.get('bonus_filtri', '')
    jarima_filtri = request.GET.get('jarima_filtri', '')

    oy_boshi, oy_oxiri = oy_oraligi(yil, oy)
    hisobot_data, jami = hisobot_data_yig(
        oy_boshi, oy_oxiri,
        filtrlar={'bonus': bonus_filtri, 'jarima': jarima_filtri}
    )
    oy_nomi = OYLAR.get(oy, '')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{oy_nomi}_{yil}"

    blue_fill = PatternFill(start_color='2001FF', end_color='2001FF', fill_type='solid')
    grey_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A1:I1')
    ws['A1'] = f"OYLIK HISOBOT — {oy_nomi} {yil}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    headers = ['№', 'Xodim', 'Bonus Ball', "Bonus Pul (so'm)", 'Jarima Ball', "Jarima Pul (so'm)", 'Jami Ball', "Jami Pul (so'm)", 'Reyting']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = blue_fill
        cell.alignment = center_align

    for idx, item in enumerate(hisobot_data, 1):
        row = idx + 3
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=f"{item['xodim'].ism} {item['xodim'].familya}")
        ws.cell(row=row, column=3, value=item['bonus_ball'])
        ws.cell(row=row, column=4, value=item['bonus_pul'])
        ws.cell(row=row, column=5, value=item['jarima_ball'])
        ws.cell(row=row, column=6, value=item['jarima_pul'])
        ws.cell(row=row, column=7, value=item['jami_ball'])
        ws.cell(row=row, column=8, value=item['jami_pul'])
        ws.cell(row=row, column=9, value=item['xodim'].reyting_ball)

    total_row = len(hisobot_data) + 4
    ws.cell(row=total_row, column=1, value='JAMI')
    ws.cell(row=total_row, column=3, value=jami['bonus_ball'])
    ws.cell(row=total_row, column=4, value=jami['bonus_pul'])
    ws.cell(row=total_row, column=5, value=jami['jarima_ball'])
    ws.cell(row=total_row, column=6, value=jami['jarima_pul'])
    ws.cell(row=total_row, column=7, value=jami['ball'])
    ws.cell(row=total_row, column=8, value=jami['pul'])

    for col in range(1, 10):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = grey_fill

    for i, width in enumerate([5, 30, 12, 15, 12, 15, 12, 15, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="oylik_hisobot_{oy_nomi}_{yil}.xlsx"'
    wb.save(response)
    return response


@staff_member_required
def barcha_malumotlar_excel(request):
    oy = int(request.GET.get('oy', timezone.now().month))
    yil = int(request.GET.get('yil', timezone.now().year))
    oy_boshi, oy_oxiri = oy_oraligi(yil, oy)
    oy_nomi = OYLAR.get(oy, '')

    wb = openpyxl.Workbook()

    blue_fill = PatternFill(start_color='2001FF', end_color='2001FF', fill_type='solid')
    green_fill = PatternFill(start_color='28a745', end_color='28a745', fill_type='solid')
    red_fill = PatternFill(start_color='dc3545', end_color='dc3545', fill_type='solid')
    orange_fill = PatternFill(start_color='fd7e14', end_color='fd7e14', fill_type='solid')
    purple_fill = PatternFill(start_color='6f42c1', end_color='6f42c1', fill_type='solid')
    grey_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    title_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )

    def style_header(ws, row, cols, fill):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = center
            cell.border = thin_border

    def style_data_row(ws, row, cols):
        for col in range(1, cols + 1):
            ws.cell(row=row, column=col).border = thin_border

    # ==================== 1. XODIMLAR (faol + arxivlangan) ====================
    ws_xodim = wb.active
    ws_xodim.title = 'Xodimlar'
    xodimlar = Xodim.objects.filter(active=True).order_by('-is_archived', '-reyting_ball')

    ws_xodim.merge_cells('A1:K1')
    ws_xodim['A1'] = f"BARCHA XODIMLAR — {oy_nomi} {yil}"
    ws_xodim['A1'].font = title_font
    ws_xodim['A1'].alignment = center

    xodim_headers = ['№', 'Ism', 'Familya', 'Telefon', 'Lavozim', 'Bonus Ball', 'Bonus Pul', 'Jarima Ball', 'Jarima Pul', 'Reyting Ball', 'Holati']
    for col, h in enumerate(xodim_headers, 1):
        ws_xodim.cell(row=3, column=col, value=h)
    style_header(ws_xodim, 3, len(xodim_headers), blue_fill)

    for idx, x in enumerate(xodimlar, 1):
        row = idx + 3
        ws_xodim.cell(row=row, column=1, value=idx)
        ws_xodim.cell(row=row, column=2, value=x.ism)
        ws_xodim.cell(row=row, column=3, value=x.familya)
        ws_xodim.cell(row=row, column=4, value=x.telefon)
        ws_xodim.cell(row=row, column=5, value=x.lavozim)
        ws_xodim.cell(row=row, column=6, value=x.bonus_ball)
        ws_xodim.cell(row=row, column=7, value=float(x.bonus_pul))
        ws_xodim.cell(row=row, column=8, value=x.jarima_ball)
        ws_xodim.cell(row=row, column=9, value=float(x.jarima_pul))
        ws_xodim.cell(row=row, column=10, value=x.reyting_ball)
        ws_xodim.cell(row=row, column=11, value='Arxivlangan' if x.is_archived else 'Faol')
        style_data_row(ws_xodim, row, len(xodim_headers))

    total_r = len(xodimlar) + 4
    ws_xodim.cell(row=total_r, column=1, value='JAMI')
    ws_xodim.cell(row=total_r, column=6, value=sum(x.bonus_ball for x in xodimlar))
    ws_xodim.cell(row=total_r, column=7, value=float(sum(x.bonus_pul for x in xodimlar)))
    ws_xodim.cell(row=total_r, column=8, value=sum(x.jarima_ball for x in xodimlar))
    ws_xodim.cell(row=total_r, column=9, value=float(sum(x.jarima_pul for x in xodimlar)))
    for col in range(1, len(xodim_headers) + 1):
        c = ws_xodim.cell(row=total_r, column=col)
        c.font = bold_font
        c.fill = grey_fill

    for i, w in enumerate([5, 15, 15, 18, 18, 12, 14, 12, 14, 12, 14], 1):
        ws_xodim.column_dimensions[get_column_letter(i)].width = w

    # ==================== 2. BONUSLAR ====================
    ws_bonus = wb.create_sheet('Bonuslar')
    bonuslar = BonusRecord.objects.filter(
        sana__date__gte=oy_boshi, sana__date__lte=oy_oxiri
    ).select_related('xodim', 'sabab', 'created_by').order_by('-sana')

    ws_bonus.merge_cells('A1:H1')
    ws_bonus['A1'] = f"BONUSLAR — {oy_nomi} {yil}"
    ws_bonus['A1'].font = title_font
    ws_bonus['A1'].alignment = center

    bonus_headers = ['№', 'Xodim', 'Sabab', 'Ball', "Pul (so'm)", 'Izoh', 'Berdi', 'Sana']
    for col, h in enumerate(bonus_headers, 1):
        ws_bonus.cell(row=3, column=col, value=h)
    style_header(ws_bonus, 3, len(bonus_headers), green_fill)

    for idx, b in enumerate(bonuslar, 1):
        row = idx + 3
        ws_bonus.cell(row=row, column=1, value=idx)
        ws_bonus.cell(row=row, column=2, value=f"{b.xodim.ism} {b.xodim.familya}")
        ws_bonus.cell(row=row, column=3, value=str(b.sabab) if b.sabab else 'Qo\'lda')
        ws_bonus.cell(row=row, column=4, value=b.ball_miqdori)
        ws_bonus.cell(row=row, column=5, value=float(b.pul_miqdori))
        ws_bonus.cell(row=row, column=6, value=b.izoh)
        ws_bonus.cell(row=row, column=7, value=b.created_by.get_full_name() if b.created_by else '')
        ws_bonus.cell(row=row, column=8, value=b.sana.strftime('%d.%m.%Y %H:%M') if b.sana else '')
        style_data_row(ws_bonus, row, len(bonus_headers))

    total_b = len(bonuslar) + 4
    ws_bonus.cell(row=total_b, column=1, value='JAMI')
    ws_bonus.cell(row=total_b, column=4, value=sum(b.ball_miqdori for b in bonuslar))
    ws_bonus.cell(row=total_b, column=5, value=float(sum(b.pul_miqdori for b in bonuslar)))
    for col in range(1, len(bonus_headers) + 1):
        c = ws_bonus.cell(row=total_b, column=col)
        c.font = bold_font
        c.fill = grey_fill

    for i, w in enumerate([5, 20, 25, 10, 14, 30, 20, 18], 1):
        ws_bonus.column_dimensions[get_column_letter(i)].width = w

    # ==================== 3. JARIMALAR ====================
    ws_jarima = wb.create_sheet('Jarimalar')
    jarimalar = JarimaRecord.objects.filter(
        sana__date__gte=oy_boshi, sana__date__lte=oy_oxiri
    ).select_related('xodim', 'sabab', 'created_by').order_by('-sana')

    ws_jarima.merge_cells('A1:H1')
    ws_jarima['A1'] = f"JARIMALAR — {oy_nomi} {yil}"
    ws_jarima['A1'].font = title_font
    ws_jarima['A1'].alignment = center

    jarima_headers = ['№', 'Xodim', 'Sabab', 'Ball', "Pul (so'm)", 'Izoh', 'Berdi', 'Sana']
    for col, h in enumerate(jarima_headers, 1):
        ws_jarima.cell(row=3, column=col, value=h)
    style_header(ws_jarima, 3, len(jarima_headers), red_fill)

    for idx, j in enumerate(jarimalar, 1):
        row = idx + 3
        ws_jarima.cell(row=row, column=1, value=idx)
        ws_jarima.cell(row=row, column=2, value=f"{j.xodim.ism} {j.xodim.familya}")
        ws_jarima.cell(row=row, column=3, value=str(j.sabab) if j.sabab else 'Qo\'lda')
        ws_jarima.cell(row=row, column=4, value=j.ball_miqdori)
        ws_jarima.cell(row=row, column=5, value=float(j.pul_miqdori))
        ws_jarima.cell(row=row, column=6, value=j.izoh)
        ws_jarima.cell(row=row, column=7, value=j.created_by.get_full_name() if j.created_by else '')
        ws_jarima.cell(row=row, column=8, value=j.sana.strftime('%d.%m.%Y %H:%M') if j.sana else '')
        style_data_row(ws_jarima, row, len(jarima_headers))

    total_j = len(jarimalar) + 4
    ws_jarima.cell(row=total_j, column=1, value='JAMI')
    ws_jarima.cell(row=total_j, column=4, value=sum(j.ball_miqdori for j in jarimalar))
    ws_jarima.cell(row=total_j, column=5, value=float(sum(j.pul_miqdori for j in jarimalar)))
    for col in range(1, len(jarima_headers) + 1):
        c = ws_jarima.cell(row=total_j, column=col)
        c.font = bold_font
        c.fill = grey_fill

    for i, w in enumerate([5, 20, 25, 10, 14, 30, 20, 18], 1):
        ws_jarima.column_dimensions[get_column_letter(i)].width = w

    # ==================== 4. MAHSULOTLAR ====================
    ws_mah = wb.create_sheet('Mahsulotlar')
    mahsulotlar = Product.objects.select_related('category').order_by('-created_at')

    ws_mah.merge_cells('A1:G1')
    ws_mah['A1'] = f"MAHSULOTLAR — Barcha ma'lumotlar"
    ws_mah['A1'].font = title_font
    ws_mah['A1'].alignment = center

    mah_headers = ['№', 'Nomi', 'Tavsifi', 'Kategoriya', "Ball", 'Omborda', 'Holati']
    for col, h in enumerate(mah_headers, 1):
        ws_mah.cell(row=3, column=col, value=h)
    style_header(ws_mah, 3, len(mah_headers), orange_fill)

    for idx, m in enumerate(mahsulotlar, 1):
        row = idx + 3
        ws_mah.cell(row=row, column=1, value=idx)
        ws_mah.cell(row=row, column=2, value=m.name)
        ws_mah.cell(row=row, column=3, value=m.description[:100] if m.description else '')
        ws_mah.cell(row=row, column=4, value=m.category.name if m.category else '')
        ws_mah.cell(row=row, column=5, value=m.price_points)
        ws_mah.cell(row=row, column=6, value=m.stock)
        ws_mah.cell(row=row, column=7, value='Faol' if m.is_active else 'Nofaol')
        style_data_row(ws_mah, row, len(mah_headers))

    for i, w in enumerate([5, 25, 30, 20, 10, 10, 10], 1):
        ws_mah.column_dimensions[get_column_letter(i)].width = w

    # ==================== 5. BUYURTMALAR ====================
    ws_buy = wb.create_sheet('Buyurtmalar')
    buyurtmalar = ProductOrder.objects.select_related('user', 'product').order_by('-created_at')

    ws_buy.merge_cells('A1:H1')
    ws_buy['A1'] = f"BUYURTMALAR — Barcha ma'lumotlar"
    ws_buy['A1'].font = title_font
    ws_buy['A1'].alignment = center

    buy_headers = ['№', 'Foydalanuvchi', 'Mahsulot', 'Ball', 'Holati', 'Yaratilgan', 'Tasdiqlangan', 'Rad etilgan']
    for col, h in enumerate(buy_headers, 1):
        ws_buy.cell(row=3, column=col, value=h)
    style_header(ws_buy, 3, len(buy_headers), purple_fill)

    status_map = {'PENDING': 'Kutilmoqda', 'APPROVED': 'Tasdiqlangan', 'REJECTED': 'Rad etilgan'}
    for idx, bo in enumerate(buyurtmalar, 1):
        row = idx + 3
        ws_buy.cell(row=row, column=1, value=idx)
        ws_buy.cell(row=row, column=2, value=bo.user.get_full_name() or bo.user.username)
        ws_buy.cell(row=row, column=3, value=bo.product.name)
        ws_buy.cell(row=row, column=4, value=bo.points_spent)
        ws_buy.cell(row=row, column=5, value=status_map.get(bo.status, bo.status))
        ws_buy.cell(row=row, column=6, value=bo.created_at.strftime('%d.%m.%Y %H:%M') if bo.created_at else '')
        ws_buy.cell(row=row, column=7, value=bo.approved_at.strftime('%d.%m.%Y %H:%M') if bo.approved_at else '')
        ws_buy.cell(row=row, column=8, value=bo.rejected_at.strftime('%d.%m.%Y %H:%M') if bo.rejected_at else '')
        style_data_row(ws_buy, row, len(buy_headers))

    for i, w in enumerate([5, 25, 25, 10, 15, 18, 18, 18], 1):
        ws_buy.column_dimensions[get_column_letter(i)].width = w

    # ==================== 6. BONUS SABABLAR ====================
    ws_bs = wb.create_sheet('BonusSabablar')
    bonus_sabablar = BonusSabab.objects.all().order_by('-ball_miqdori')

    ws_bs.merge_cells('A1:E1')
    ws_bs['A1'] = 'BONUS SABABLAR'
    ws_bs['A1'].font = title_font
    ws_bs['A1'].alignment = center

    bs_headers = ['№', 'Nomi', "Pul (so'm)", 'Ball', 'Holati']
    for col, h in enumerate(bs_headers, 1):
        ws_bs.cell(row=3, column=col, value=h)
    style_header(ws_bs, 3, len(bs_headers), green_fill)

    for idx, s in enumerate(bonus_sabablar, 1):
        row = idx + 3
        ws_bs.cell(row=row, column=1, value=idx)
        ws_bs.cell(row=row, column=2, value=s.nom)
        ws_bs.cell(row=row, column=3, value=float(s.pul_miqdori))
        ws_bs.cell(row=row, column=4, value=s.ball_miqdori)
        ws_bs.cell(row=row, column=5, value='Faol' if s.active else 'Nofaol')
        style_data_row(ws_bs, row, len(bs_headers))

    for i, w in enumerate([5, 30, 14, 10, 10], 1):
        ws_bs.column_dimensions[get_column_letter(i)].width = w

    # ==================== 7. JARIMA SABABLAR ====================
    ws_js = wb.create_sheet('JarimaSabablar')
    jarima_sabablar = JarimaSabab.objects.all().order_by('-ball_miqdori')

    ws_js.merge_cells('A1:E1')
    ws_js['A1'] = 'JARIMA SABABLAR'
    ws_js['A1'].font = title_font
    ws_js['A1'].alignment = center

    js_headers = ['№', 'Nomi', "Pul (so'm)", 'Ball', 'Holati']
    for col, h in enumerate(js_headers, 1):
        ws_js.cell(row=3, column=col, value=h)
    style_header(ws_js, 3, len(js_headers), red_fill)

    for idx, s in enumerate(jarima_sabablar, 1):
        row = idx + 3
        ws_js.cell(row=row, column=1, value=idx)
        ws_js.cell(row=row, column=2, value=s.nom)
        ws_js.cell(row=row, column=3, value=float(s.pul_miqdori))
        ws_js.cell(row=row, column=4, value=s.ball_miqdori)
        ws_js.cell(row=row, column=5, value='Faol' if s.active else 'Nofaol')
        style_data_row(ws_js, row, len(js_headers))

    for i, w in enumerate([5, 30, 14, 10, 10], 1):
        ws_js.column_dimensions[get_column_letter(i)].width = w

    # ==================== 8. OYLIK HISOBOT ====================
    ws_oylik = wb.create_sheet(f'{oy_nomi}_{yil}')
    hisobot_data, jami = hisobot_data_yig(oy_boshi, oy_oxiri)

    ws_oylik.merge_cells('A1:I1')
    ws_oylik['A1'] = f"OYLIK HISOBOT — {oy_nomi} {yil}"
    ws_oylik['A1'].font = title_font
    ws_oylik['A1'].alignment = center

    oylik_headers = ['№', 'Xodim', 'Bonus Ball', "Bonus Pul (so'm)", 'Jarima Ball', "Jarima Pul (so'm)", 'Jami Ball', "Jami Pul (so'm)", 'Reyting']
    for col, h in enumerate(oylik_headers, 1):
        ws_oylik.cell(row=3, column=col, value=h)
    style_header(ws_oylik, 3, len(oylik_headers), blue_fill)

    for idx, item in enumerate(hisobot_data, 1):
        row = idx + 3
        ws_oylik.cell(row=row, column=1, value=idx)
        ws_oylik.cell(row=row, column=2, value=f"{item['xodim'].ism} {item['xodim'].familya}")
        ws_oylik.cell(row=row, column=3, value=item['bonus_ball'])
        ws_oylik.cell(row=row, column=4, value=item['bonus_pul'])
        ws_oylik.cell(row=row, column=5, value=item['jarima_ball'])
        ws_oylik.cell(row=row, column=6, value=item['jarima_pul'])
        ws_oylik.cell(row=row, column=7, value=item['jami_ball'])
        ws_oylik.cell(row=row, column=8, value=item['jami_pul'])
        ws_oylik.cell(row=row, column=9, value=item['xodim'].reyting_ball)
        style_data_row(ws_oylik, row, len(oylik_headers))

    total_y = len(hisobot_data) + 4
    ws_oylik.cell(row=total_y, column=1, value='JAMI')
    ws_oylik.cell(row=total_y, column=3, value=jami['bonus_ball'])
    ws_oylik.cell(row=total_y, column=4, value=jami['bonus_pul'])
    ws_oylik.cell(row=total_y, column=5, value=jami['jarima_ball'])
    ws_oylik.cell(row=total_y, column=6, value=jami['jarima_pul'])
    ws_oylik.cell(row=total_y, column=7, value=jami['ball'])
    ws_oylik.cell(row=total_y, column=8, value=jami['pul'])
    for col in range(1, len(oylik_headers) + 1):
        c = ws_oylik.cell(row=total_y, column=col)
        c.font = bold_font
        c.fill = grey_fill

    for i, w in enumerate([5, 30, 12, 15, 12, 15, 12, 15, 12], 1):
        ws_oylik.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="barcha_malumotlar_{oy_nomi}_{yil}.xlsx"'
    wb.save(response)
    return response


# ============================================================
# PROFIL
# ============================================================

@login_required
def mening_profilim(request):
    """Xodimning shaxsiy profil sahifasi"""
    try:
        xodim = request.user.xodim
    except Exception:
        messages.error(request, "Siz xodim sifatida ro'yxatdan o'tmagansiz!")
        return redirect('dashboard')

    if request.method == 'POST':
        if 'rasm_yuklash' in request.POST:
            form = XodimRasmForm(request.POST, request.FILES, instance=xodim)
            if form.is_valid():
                form.save()
                messages.success(request, 'Rasm yuklandi!')
                return redirect('mening_profilim')
            else:
                messages.error(request, 'Rasm yuklashda xatolik yuz berdi!')
                return redirect('mening_profilim')
        elif 'rasm_ochirish' in request.POST:
            if xodim.rasm:
                xodim.rasm.delete()
                xodim.rasm = None
                xodim.save()
                messages.success(request, "Rasm o'chirildi!")
            else:
                messages.error(request, "Rasm mavjud emas!")
            return redirect('mening_profilim')

    # Bonus va jarimalarni olish
    bonuslar = BonusRecord.objects.filter(xodim=xodim).select_related('sabab').order_by('-sana')[:50]
    jarimalar = JarimaRecord.objects.filter(xodim=xodim).select_related('sabab').order_by('-sana')[:50]
    
    # Xaridlar tarixi
    xaridlar = ProductOrder.objects.filter(user=request.user).select_related('product').order_by('-created_at')[:50]
    
    # Reytingdagi o'rni
    joylashuv = Xodim.objects.filter(reyting_ball__gt=xodim.reyting_ball, is_archived=False).count() + 1
    
    # Jami xodimlar soni
    jami_xodimlar = Xodim.objects.filter(active=True, is_archived=False).count()
    
    # Yechilgan pullarni hisoblash (agar modelda bo'lmasa)
    bonus_ball_yechilgan = getattr(xodim, 'bonus_ball_yechilgan', 0)
    bonus_pul_yechilgan = getattr(xodim, 'bonus_pul_yechilgan', 0)
    jarima_ball_yechilgan = getattr(xodim, 'jarima_ball_yechilgan', 0)
    jarima_pul_yechilgan = getattr(xodim, 'jarima_pul_yechilgan', 0)
    
    # Jami bonus va jarima (qolgan) - xarid_ball ni ham hisobga olish
    xarid_ball = getattr(xodim, 'xarid_ball', 0)
    jami_bonus_ball = xodim.bonus_ball - bonus_ball_yechilgan - xarid_ball
    jami_bonus_pul = xodim.bonus_pul - bonus_pul_yechilgan
    jami_jarima_ball = xodim.jarima_ball - jarima_ball_yechilgan
    jami_jarima_pul = xodim.jarima_pul - jarima_pul_yechilgan
    
    # Sof reyting
    sof_reyting_ball = jami_bonus_ball - jami_jarima_ball
    sof_reyting_pul = jami_bonus_pul - jami_jarima_pul

    context = {
        'xodim': xodim,
        'bonuslar': bonuslar,
        'jarimalar': jarimalar,
        'xaridlar': xaridlar,
        'joylashuv': joylashuv,
        'jami_xodimlar': jami_xodimlar,
        
        # Umumiy qiymatlar
        'umumiy_bonus_ball': xodim.bonus_ball,
        'umumiy_bonus_pul': xodim.bonus_pul,
        'umumiy_jarima_ball': xodim.jarima_ball,
        'umumiy_jarima_pul': xodim.jarima_pul,
        
        # Yechilgan qiymatlar
        'bonus_ball_yechilgan': bonus_ball_yechilgan,
        'bonus_pul_yechilgan': bonus_pul_yechilgan,
        'jarima_ball_yechilgan': jarima_ball_yechilgan,
        'jarima_pul_yechilgan': jarima_pul_yechilgan,
        
        # Qolgan (jami) qiymatlar
        'jami_bonus_ball': jami_bonus_ball,
        'jami_bonus_pul': jami_bonus_pul,
        'jami_jarima_ball': jami_jarima_ball,
        'jami_jarima_pul': jami_jarima_pul,
        
        # Sof reyting
        'reyting_ball': xodim.reyting_ball,
        'reyting_pul': xodim.reyting_pul,
        'sof_reyting_ball': sof_reyting_ball,
        'sof_reyting_pul': sof_reyting_pul,
    }
    
    return render(request, 'main/mening_profilim.html', context)
# main/views.py

@login_required
def profil_sozlamalari(request):
    """Profil sozlamalari sahifasi"""
    user = request.user
    xodim = getattr(user, 'xodim', None)

    login_form = UserEditForm(instance=user)
    password_form = PasswordChangeForm(user)

    if request.method == 'POST':
        if 'login_form' in request.POST:
            login_form = UserEditForm(request.POST, instance=user)
            if login_form.is_valid():
                login_form.save()
                messages.success(request, "Login o'zgartirildi!")
                return redirect('profil_sozlamalari')
            else:
                messages.error(request, "Xatolik yuz berdi!")
        elif 'password_form' in request.POST:
            password_form = PasswordChangeForm(user, request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)
                messages.success(request, "Parol o'zgartirildi!")
                return redirect('profil_sozlamalari')
            else:
                messages.error(request, "Parol xatolik!")

    return render(request, 'main/profil_sozlamalari.html', {
        'user': user,
        'xodim': xodim,
        'login_form': login_form,
        'password_form': password_form,
    })


@login_required
def login_ozgartirish(request):
    """Login o'zgartirish sahifasi"""
    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Login o'zgartirildi!")
            return redirect('profil_sozlamalari')
        else:
            messages.error(request, "Xatolik yuz berdi!")
    else:
        form = UserEditForm(instance=request.user)
    
    return render(request, 'main/login_ozgartirish.html', {
        'form': form,
        'title': "Login o'zgartirish"
    })


@login_required
def parol_ozgartirish(request):
    """Parol o'zgartirish sahifasi"""
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, "Parol o'zgartirildi!")
            return redirect('profil_sozlamalari')
        else:
            messages.error(request, "Parol xatolik!")
    else:
        form = PasswordChangeForm(request.user)
    
    return render(request, 'main/parol_ozgartirish.html', {
        'form': form,
        'title': "Parol o'zgartirish"
    })


@login_required
def rasm_ochirish(request):
    """Rasm o'chirish"""
    try:
        xodim = request.user.xodim
        if xodim.rasm:
            xodim.rasm.delete()
            xodim.rasm = None
            xodim.save()
            messages.success(request, "Rasm o'chirildi!")
    except Exception:
        messages.error(request, 'Xatolik yuz berdi!')
    return redirect('mening_profilim')
@login_required
def profil_sozlamalari(request):
    user = request.user
    xodim = getattr(user, 'xodim', None)

    login_form = UserEditForm(instance=user)
    password_form = PasswordChangeForm(user)

    if request.method == 'POST':
        if 'login_form' in request.POST:
            login_form = UserEditForm(request.POST, instance=user)
            if login_form.is_valid():
                login_form.save()
                messages.success(request, "Login o'zgartirildi!")
                return redirect('profil_sozlamalari')
        elif 'password_form' in request.POST:
            password_form = PasswordChangeForm(user, request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)
                messages.success(request, "Parol o'zgartirildi!")
                return redirect('profil_sozlamalari')

    return render(request, 'main/profil_sozlamalari.html', {
        'user': user,
        'xodim': xodim,
        'login_form': login_form,
        'password_form': password_form,
    })


# ============================================================
# SABABLAR BOSHQARUVI
# ============================================================

@staff_member_required
def sabablar_boshqaruvi(request):
    if request.method == 'POST':
        if 'bonus_qoshish' in request.POST:
            BonusSabab.objects.create(
                nom=request.POST.get('bonus_nom'),
                pul_miqdori=request.POST.get('bonus_pul'),
                ball_miqdori=request.POST.get('bonus_ball'),
                active=True
            )
            messages.success(request, "Bonus sababi qo'shildi!")
        elif 'jarima_qoshish' in request.POST:
            JarimaSabab.objects.create(
                nom=request.POST.get('jarima_nom'),
                pul_miqdori=request.POST.get('jarima_pul'),
                ball_miqdori=request.POST.get('jarima_ball'),
                active=True
            )
            messages.success(request, "Jarima sababi qo'shildi!")
        elif 'ochirish' in request.POST:
            tur = request.POST.get('tur')
            pk = request.POST.get('pk')
            if tur == 'bonus':
                BonusSabab.objects.filter(id=pk).delete()
            else:
                JarimaSabab.objects.filter(id=pk).delete()
            messages.success(request, "Sabab o'chirildi!")
        return redirect('sabablar_boshqaruvi')

    return render(request, 'main/sabablar_boshqaruvi.html', {
        'bonus_sabablar': BonusSabab.objects.all().order_by('-ball_miqdori'),
        'jarima_sabablar': JarimaSabab.objects.all().order_by('-ball_miqdori'),
    })


# ============================================================
# TARIXLAR
# ============================================================

@login_required
def tarixlar(request, pk=None):
    if pk:
        xodim = get_object_or_404(Xodim, pk=pk)
        return render(request, 'main/tarixlar.html', {
            'xodim': xodim,
            'bonuslar': BonusRecord.objects.filter(xodim=xodim).select_related('sabab').order_by('-sana'),
            'jarimalar': JarimaRecord.objects.filter(xodim=xodim).select_related('sabab').order_by('-sana'),
            'ozgarishlar': OzgartirishTarixi.objects.filter(xodim=xodim).order_by('-sana'),
            'yagona_xodim': True,
        })

    if not request.user.is_staff:
        return HttpResponseForbidden("Sizda bu sahifaga kirish ruxsati yo'q!")

    qidiruv = request.GET.get('qidiruv', '')

    def qidirish(Model, extra_q, qidiruv):
        base = Q(xodim__ism__icontains=qidiruv) | Q(xodim__familya__icontains=qidiruv) | extra_q
        return Model.objects.filter(base) if qidiruv else Model.objects.all()

    bonuslar = qidirish(BonusRecord, Q(izoh__icontains=qidiruv), qidiruv).select_related('xodim', 'sabab').order_by('-sana')[:100]
    jarimalar = qidirish(JarimaRecord, Q(izoh__icontains=qidiruv), qidiruv).select_related('xodim', 'sabab').order_by('-sana')[:100]
    ozgarishlar = qidirish(OzgartirishTarixi, Q(sabab__icontains=qidiruv), qidiruv).select_related('xodim', 'admin').order_by('-sana')[:100]

    return render(request, 'main/tarixlar.html', {
        'bonuslar': bonuslar,
        'jarimalar': jarimalar,
        'ozgarishlar': ozgarishlar,
        'yagona_xodim': False,
        'qidiruv': qidiruv,
    })


# ============================================================
# ADMIN DASHBOARD
# ============================================================

@staff_member_required
def admin_dashboard(request):
    bugun = timezone.now().date()
    pending_orders = ProductOrder.objects.filter(status='PENDING').count()
    
    oxirgi_bonuslar = BonusRecord.objects.select_related('xodim').order_by('-sana')[:20]
    oxirgi_jarimalar = JarimaRecord.objects.select_related('xodim').order_by('-sana')[:20]
    
    bon_agg = BonusRecord.objects.aggregate(
        bball=Sum('ball_miqdori'), bpul=Sum('pul_miqdori')
    )
    jar_agg = JarimaRecord.objects.aggregate(
        jball=Sum('ball_miqdori'), jpul=Sum('pul_miqdori')
    )
    umumiy_bonus_ball = bon_agg['bball'] or 0
    umumiy_bonus_pul = bon_agg['bpul'] or 0
    umumiy_jarima_ball = jar_agg['jball'] or 0
    umumiy_jarima_pul = jar_agg['jpul'] or 0
    
    xodimlar_soni = Xodim.objects.count()
    bugun_bonus = BonusRecord.objects.filter(sana__date=bugun).count()
    bugun_jarima = JarimaRecord.objects.filter(sana__date=bugun).count()
    
    return render(request, 'main/admin_dashboard.html', {
        'jami_xodimlar': xodimlar_soni,
        'jami_bonuslar': BonusRecord.objects.count(),
        'jami_jarimalar': JarimaRecord.objects.count(),
        'umumiy_ball': Xodim.objects.aggregate(Sum('reyting_ball'))['reyting_ball__sum'] or 0,
        'umumiy_pul': Xodim.objects.aggregate(Sum('reyting_pul'))['reyting_pul__sum'] or 0,
        'bugun_bonus': bugun_bonus,
        'bugun_jarima': bugun_jarima,
        'top_xodimlar': Xodim.objects.filter(active=True, is_archived=False).order_by('-reyting_ball')[:10],
        'pending_orders': pending_orders,
        'oxirgi_bonuslar': oxirgi_bonuslar,
        'oxirgi_jarimalar': oxirgi_jarimalar,
        'bugun_qoshilgan': bugun_bonus + bugun_jarima,
        'faol_sabablar': BonusSabab.objects.count() + JarimaSabab.objects.count(),
        'umumiy_bonus_ball': umumiy_bonus_ball,
        'umumiy_bonus_pul': umumiy_bonus_pul,
        'umumiy_jarima_ball': umumiy_jarima_ball,
        'umumiy_jarima_pul': umumiy_jarima_pul,
        'xodimlar_soni': xodimlar_soni,
    })


def admin(request):
    return render(request, 'main/admin.html')











# main/views.py

@staff_member_required
def xodim_ballarni_tekislash(request, pk):
    """Xodimning bonus va jarima ballarini to'g'ridan-to'g'ri tahrirlash"""
    xodim = get_object_or_404(Xodim, pk=pk)
    
    if request.method == 'POST':
        sabab = request.POST.get('sabab', '').strip()
        if not sabab:
            messages.error(request, "O'zgartirish sababini yozishingiz kerak!")
            return redirect('xodim_ballarni_tekislash', pk=xodim.pk)
        
        try:
            bonus_ball = int(request.POST.get('bonus_ball', 0))
            bonus_pul = Decimal(str(float(request.POST.get('bonus_pul', 0))))
            jarima_ball = int(request.POST.get('jarima_ball', 0))
            jarima_pul = Decimal(str(float(request.POST.get('jarima_pul', 0))))
        except (ValueError, TypeError):
            messages.error(request, "Noto'g'ri qiymat kiritildi!")
            return redirect('xodim_ballarni_tekislash', pk=xodim.pk)
        
        # Eski qiymatlarni saqlash
        eski_bonus_ball = xodim.bonus_ball
        eski_bonus_pul = xodim.bonus_pul
        eski_jarima_ball = xodim.jarima_ball
        eski_jarima_pul = xodim.jarima_pul
        
        # Yangi qiymatlarni o'rnatish
        xodim.bonus_ball = bonus_ball
        xodim.bonus_pul = bonus_pul
        xodim.jarima_ball = jarima_ball
        xodim.jarima_pul = jarima_pul
        
        # Reytingni qayta hisoblash
        xodim.reyting_ball = bonus_ball - jarima_ball
        xodim.reyting_pul = (bonus_pul - xodim.bonus_pul_yechilgan) - (jarima_pul - xodim.jarima_pul_yechilgan)
        xodim.save()
        
        # Tarixga yozish
        OzgartirishTarixi.objects.create(
            xodim=xodim,
            admin=request.user,
            sabab=f"Ballarni tekislash. Sabab: {sabab}",
            eski_bonus_ball=eski_bonus_ball,
            eski_bonus_pul=eski_bonus_pul,
            eski_jarima_ball=eski_jarima_ball,
            eski_jarima_pul=eski_jarima_pul,
            yangi_bonus_ball=bonus_ball,
            yangi_bonus_pul=bonus_pul,
            yangi_jarima_ball=jarima_ball,
            yangi_jarima_pul=jarima_pul,
        )
        
        messages.success(request, f"{xodim.ism} {xodim.familya} ballari yangilandi!")
        return redirect('xodim_detail', pk=xodim.pk)
    
    return render(request, 'main/xodim_ballarni_tekislash.html', {'xodim': xodim})





# main/views.py
@staff_member_required
def bonus_ball_yechish(request, pk):
    """Bonus ballni yechish"""
    xodim = get_object_or_404(Xodim, pk=pk)
    
    if request.method == 'POST':
        try:
            yechiladigan_ball = int(request.POST.get('ball', 0))
            sabab = request.POST.get('sabab', '').strip()
            
            if yechiladigan_ball <= 0:
                messages.error(request, "Ball miqdori 0 dan katta bo'lishi kerak!")
                return redirect('bonus_ball_yechish', pk=xodim.pk)
            
            if not sabab:
                messages.error(request, "Sababni yozishingiz kerak!")
                return redirect('bonus_ball_yechish', pk=xodim.pk)
            
            if yechiladigan_ball > xodim.jami_bonus_ball:
                messages.error(request, f"Yetarli bonus ball mavjud emas! Mavjud: {xodim.jami_bonus_ball} ball")
                return redirect('bonus_ball_yechish', pk=xodim.pk)
            
            # Yechish
            xodim.bonus_ball_yechilgan += yechiladigan_ball
            xodim.reyting_ball = xodim.jami_bonus_ball - xodim.jami_jarima_ball
            xodim.save(update_fields=['bonus_ball_yechilgan', 'reyting_ball'])
            
            # TO'G'RI - FAQAT MAVJUD MAYDONLAR
            OzgartirishTarixi.objects.create(
                xodim=xodim,
                admin=request.user,
                sabab=f"Bonus balldan {yechiladigan_ball} ball yechildi. Sabab: {sabab}",
            )
            
            messages.success(request, f"✅ {yechiladigan_ball} ball bonus balldan yechildi! Qolgan: {xodim.jami_bonus_ball} ball")
            return redirect('xodim_detail', pk=xodim.pk)
            
        except (ValueError, TypeError) as e:
            messages.error(request, f"Xatolik: {e}")
            return redirect('bonus_ball_yechish', pk=xodim.pk)
    
    return render(request, 'main/ball_yechish.html', {
        'xodim': xodim,
        'tur': 'bonus_ball',
        'umumiy': xodim.bonus_ball,
        'yechilgan': xodim.bonus_ball_yechilgan,
        'qoldiq': xodim.jami_bonus_ball,
    })

@staff_member_required
def jarima_ball_yechish(request, pk):
    """Jarima ballni yechish"""
    xodim = get_object_or_404(Xodim, pk=pk)
    
    if request.method == 'POST':
        try:
            yechiladigan_ball = int(request.POST.get('ball', 0))
            sabab = request.POST.get('sabab', '').strip()
            
            if yechiladigan_ball <= 0:
                messages.error(request, "Ball miqdori 0 dan katta bo'lishi kerak!")
                return redirect('jarima_ball_yechish', pk=xodim.pk)
            
            if not sabab:
                messages.error(request, "Sababni yozishingiz kerak!")
                return redirect('jarima_ball_yechish', pk=xodim.pk)
            
            if yechiladigan_ball > xodim.jami_jarima_ball:
                messages.error(request, f"Yetarli jarima ball mavjud emas! Mavjud: {xodim.jami_jarima_ball} ball")
                return redirect('jarima_ball_yechish', pk=xodim.pk)
            
            # Yechish
            xodim.jarima_ball_yechish(yechiladigan_ball)
            
            # Tarixga yozish
            OzgartirishTarixi.objects.create(
                xodim=xodim,
                admin=request.user,
                sabab=f"Jarima balldan {yechiladigan_ball} ball yechildi. Sabab: {sabab}",
            )
            
            messages.success(request, f"✅ {yechiladigan_ball} ball jarima balldan yechildi! Qolgan: {xodim.jami_jarima_ball} ball")
            return redirect('xodim_detail', pk=xodim.pk)
            
        except (ValueError, TypeError):
            messages.error(request, "Noto'g'ri ball miqdori!")
            return redirect('jarima_ball_yechish', pk=xodim.pk)
    
    return render(request, 'main/ball_yechish.html', {
        'xodim': xodim,
        'tur': 'jarima_ball',
        'umumiy': xodim.jarima_ball,
        'yechilgan': xodim.jarima_ball_yechilgan,
        'qoldiq': xodim.jami_jarima_ball,
    })


# ============================================================
# MAHSULOT QOSHISH
# ============================================================

@staff_member_required
def mahsulot_qoshish(request):
    from .forms import ProductForm
    categories = Category.objects.all().order_by('order', 'name')
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save()
            # Telegram xabar
            from .services import send_telegram_message
            send_telegram_message(
                f"🛍 <b>YANGI MAHSULOT</b>\n\n"
                f"📦 Nomi: {product.name}\n"
                f"⭐ Ball: {product.price_points}\n"
                f"📦 Ombor: {product.stock} ta",
                thread_id=None
            )
            # Notification for all users
            xodimlar = Xodim.objects.filter(active=True, is_archived=False)
            for xodim in xodimlar:
                send_notification(
                    xodim.user,
                    "Yangi mahsulot",
                    f"\"{product.name}\" mahsuloti qo'shildi! {product.price_points} ball",
                    '/shop/'
                )
            messages.success(request, "✅ Mahsulot muvaffaqiyatli qo'shildi!")
            return redirect('admin_dashboard')
        else:
            messages.error(request, "Xatolik yuz berdi!")
    else:
        form = ProductForm()
    return render(request, 'main/mahsulot_form.html', {'form': form, 'categories': categories})


@login_required
def shop_edit_view(request, product_id):
    from .forms import ProductForm
    categories = Category.objects.all().order_by('order', 'name')
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Mahsulot tahrirlandi!")
            return redirect('admin_mahsulotlar')
        else:
            messages.error(request, "Xatolik yuz berdi!")
    else:
        form = ProductForm(instance=product)
    return render(request, 'main/mahsulot_form.html', {'form': form, 'edit': True, 'categories': categories})


@staff_member_required
def admin_mahsulotlar(request):
    from django.db.models import Count
    products = Product.objects.all().select_related('category').order_by('-created_at')
    categories = Category.objects.annotate(product_count=Count('products')).order_by('order', 'name')
    total_count = products.count()
    coming_soon_count = products.filter(is_coming_soon=True).count()
    return render(request, 'main/admin_mahsulotlar.html', {
        'products': products,
        'categories': categories,
        'total_count': total_count,
        'coming_soon_count': coming_soon_count,
    })


@staff_member_required
def mahsulot_ochirish(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        product.delete()
        messages.success(request, "❌ Mahsulot o'chirildi!")
        return redirect('admin_mahsulotlar')
    return render(request, 'main/mahsulot_ochirish.html', {'product': product})


@staff_member_required
def kategoriya_qoshish(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        order = request.POST.get('order', 0)
        if name:
            Category.objects.create(name=name, order=order)
            messages.success(request, f"✅ '{name}' kategoriyasi qo'shildi!")
        else:
            messages.error(request, "Kategoriya nomini kiriting!")
        return redirect('admin_mahsulotlar')
    return redirect('admin_mahsulotlar')


@staff_member_required
def kategoriya_ochirish(request, pk):
    if request.method == 'POST':
        category = get_object_or_404(Category, pk=pk)
        name = category.name
        category.delete()
        messages.success(request, f"❌ '{name}' kategoriyasi o'chirildi!")
    return redirect('admin_mahsulotlar')


# ============================================================
# SHOP
# ============================================================

@login_required
def shop_view(request):
    categories = Category.objects.all().order_by('order', 'name')
    products = Product.objects.filter(is_active=True).select_related('category')
    try:
        xodim = request.user.xodim
        available_ball = get_available_shop_ball(xodim)
    except Exception:
        xodim = None
        available_ball = 0

    # Guruhlash: kategoriya bo'yicha
    categorized = []
    uncategorized = products.filter(category__isnull=True)
    for cat in categories:
        cat_products = products.filter(category=cat)
        if cat_products.exists():
            categorized.append({
                'category': cat,
                'products': cat_products
            })
    if uncategorized.exists():
        categorized.append({
            'category': None,
            'products': uncategorized
        })

    return render(request, 'main/shop.html', {
        'categorized': categorized,
        'products': products,
        'xodim': xodim,
        'available_ball': available_ball,
    })


@login_required
def purchase_view(request, product_id):
    if request.method != 'POST':
        return redirect('shop')

    try:
        xodim = request.user.xodim
        order = purchase_product(request.user, product_id)
        messages.success(
            request,
            f"✅ Buyurtma yaratildi! Mahsulot: {order.product.name}. "
            f"Admin tasdiqlashini kuting."
        )
    except Xodim.DoesNotExist:
        messages.error(request, "Siz xodim sifatida ro'yxatdan o'tmagansiz!")
    except ValueError as e:
        messages.error(request, str(e))
    except Product.DoesNotExist:
        messages.error(request, "Mahsulot topilmadi!")
    except Exception as e:
        messages.error(request, f"Xatolik yuz berdi: {e}")

    return redirect('shop')


# ============================================================
# MY ORDERS
# ============================================================

@login_required
def my_orders_view(request):
    orders = ProductOrder.objects.filter(user=request.user).select_related('product')

    status_colors = {
        'PENDING': 'bg-yellow-100 text-yellow-800',
        'APPROVED': 'bg-green-100 text-green-800',
        'REJECTED': 'bg-red-100 text-red-800',
    }

    return render(request, 'main/my_orders.html', {
        'orders': orders,
        'status_colors': status_colors,
    })


# ============================================================
# ADMIN ORDERS
# ============================================================

@staff_member_required
def admin_order_list(request):
    status_filter = request.GET.get('status', '')
    orders = ProductOrder.objects.all().select_related('user', 'product')

    if status_filter:
        orders = orders.filter(status=status_filter)

    orders = orders.order_by('-created_at')

    return render(request, 'main/admin_orders.html', {
        'orders': orders,
        'status_filter': status_filter,
    })


@staff_member_required
def admin_order_approve(request, order_id):
    if request.method != 'POST':
        return redirect('admin_order_list')

    try:
        order = approve_order(order_id)
        messages.success(request, f"✅ Buyurtma #{order.id} tasdiqlandi!")
    except ValueError as e:
        messages.error(request, str(e))
    except ProductOrder.DoesNotExist:
        messages.error(request, "Buyurtma topilmadi!")
    except Exception as e:
        messages.error(request, f"Xatolik: {e}")

    return redirect('admin_order_list')


@staff_member_required
def admin_order_reject(request, order_id):
    order = get_object_or_404(ProductOrder, id=order_id)

    if request.method == 'POST':
        form = OrderRejectForm(request.POST)
        if form.is_valid():
            try:
                reject_order(order_id, form.cleaned_data['reject_reason'])
                messages.success(request, f"❌ Buyurtma #{order.id} rad etildi!")
                return redirect('admin_order_list')
            except ValueError as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, f"Xatolik: {e}")
        else:
            messages.error(request, "Rad etish sababini yozing!")
        return redirect('admin_order_reject', order_id=order_id)

    form = OrderRejectForm()
    return render(request, 'main/admin_order_reject.html', {
        'form': form,
        'order': order,
    })


# ============================================================
# BILDIRISHNOMALAR
# ============================================================

@login_required
def notification_list(request):
    notifications = Notification.objects.filter(user=request.user)
    if request.GET.get('read') == '0':
        notifications = notifications.filter(is_read=False)
    return render(request, 'main/bildirishnomalar.html', {
        'notifications': notifications,
    })


@login_required
def notification_unread_count(request):
    from django.http import JsonResponse
    count = Notification.objects.filter(user=request.user, is_read=False).count()
    return JsonResponse({'count': count})


@login_required
def notification_mark_read(request, pk):
    notification = get_object_or_404(Notification, pk=pk, user=request.user)
    notification.is_read = True
    notification.save(update_fields=['is_read'])
    if notification.url:
        return redirect(notification.url)
    return redirect('notification_list')


@login_required
def notification_mark_all_read(request):
    Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
    messages.success(request, "Barcha bildirishnomalar o'qildi!")
    return redirect('notification_list')


# ============================================================
# PUSH BILDIRISHNOMALAR (PWA)
# ============================================================

@login_required
@csrf_exempt
def push_subscribe(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST talab qilinadi'}, status=405)

    try:
        data = json.loads(request.body)
        endpoint = data.get('endpoint', '')
        keys = data.get('keys', {})
        p256dh = keys.get('p256dh', '')
        auth = keys.get('auth', '')

        if not endpoint or not p256dh or not auth:
            return JsonResponse({'error': 'endpoint, p256dh va auth talab qilinadi'}, status=400)

        # Mavjud obunani yangilash yoki yangisini yaratish
        sub, created = PushSubscription.objects.update_or_create(
            user=request.user,
            endpoint=endpoint,
            defaults={
                'p256dh_key': p256dh,
                'auth_key': auth,
                'user_agent': request.META.get('HTTP_USER_AGENT', ''),
            }
        )

        return JsonResponse({'status': 'ok', 'created': created})
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Noto\'g\'ri JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@csrf_exempt
def push_unsubscribe(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST talab qilinadi'}, status=405)

    try:
        data = json.loads(request.body)
        endpoint = data.get('endpoint', '')
        PushSubscription.objects.filter(user=request.user, endpoint=endpoint).delete()
        return JsonResponse({'status': 'ok'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


def vapid_public_key(request):
    from django.conf import settings
    # Return the raw 65-byte key as URL-safe base64 for JS
    from cryptography.hazmat.primitives import serialization
    from cryptography.hazmat.primitives.asymmetric import ec
    from cryptography.hazmat.backends import default_backend

    try:
        pub_key = serialization.load_pem_public_key(
            settings.VAPID_PUBLIC_KEY.encode(),
            backend=default_backend()
        )
        raw_bytes = pub_key.public_bytes(
            encoding=serialization.Encoding.X962,
            format=serialization.PublicFormat.UncompressedPoint
        )
        import base64
        b64_key = base64.urlsafe_b64encode(raw_bytes).rstrip(b'=').decode()
        return JsonResponse({'public_key': b64_key})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ============================================================
# DB EXPORT - PostgreSQL dan SQLite yuklab olish
# ============================================================

@staff_member_required
def export_sqlite_view(request):
    import tempfile
    import subprocess
    import sys

    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'db_backup_{timestamp}.sqlite3'
        temp_dir = tempfile.mkdtemp()
        output_path = os.path.join(temp_dir, filename)

        python_exe = sys.executable
        manage_py = os.path.join(settings.BASE_DIR, 'manage.py')
        result = subprocess.run(
            [python_exe, manage_py, 'export_sqlite', '--output', output_path],
            capture_output=True, text=True, cwd=str(settings.BASE_DIR), timeout=120
        )

        if result.returncode != 0:
            return HttpResponse(f'Xatolik: {result.stderr}', status=500)

        with open(output_path, 'rb') as f:
            content = f.read()

        response = HttpResponse(content, content_type='application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = len(content)

        try:
            os.remove(output_path)
            os.rmdir(temp_dir)
        except:
            pass

        return response

    except subprocess.TimeoutExpired:
        return HttpResponse('Xatolik: Vaqt tugadi (120s)', status=500)
    except Exception as e:
        return HttpResponse(f'Xatolik: {str(e)}', status=500)